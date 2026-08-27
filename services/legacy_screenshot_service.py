"""
Legacy Screenshot Service - Mô phỏng phương pháp chụp ảnh từ phiên bản cũ

Port từ SosanhCTTT_phienban6_list_28.08.2025.py
Phương pháp này:
1. Mở Excel hiện hữu trên màn hình
2. Chụp ảnh trực tiếp từ màn hình (không qua temp workbook)
3. Hỗ trợ nhiều chế độ màn hình (PC, VPS, Monitor phụ)
"""

import os
import time
import glob
import utils
import config

try:
    import pythoncom
    import win32com.client
    PYWIN32_AVAILABLE = True
except ImportError:
    PYWIN32_AVAILABLE = False

try:
    from pywinauto import Application
    PYWINAUTO_AVAILABLE = True
except ImportError:
    PYWINAUTO_AVAILABLE = False

try:
    import pyautogui
    from PIL import ImageGrab, Image, ImageChops, ImageDraw
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as xlImage
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class LegacyScreenshotService:
    """
    Service chụp ảnh theo phương pháp Legacy.
    
    Khác với phương pháp hiện tại (copy vào temp workbook), 
    phương pháp này mở trực tiếp file Excel và chụp ảnh màn hình.
    """
    
    def __init__(self):
        self.saved_col_widths = {}  # Lưu độ rộng cột từ file mới
        self.screen_mode = 'pc'  # pc, vps, monitor
        self.zoom_level = 46
        self.goto_address = 'EX1'
        self.highlight_fill_color = "#ff0000"
        self.highlight_fill_opacity = 30
    
    def set_screen_mode(self, mode):
        """Đặt chế độ màn hình: 'pc', 'vps', 'monitor'"""
        self.screen_mode = mode
    
    def set_goto_address(self, address):
        """Đặt địa chỉ cell để định vị trước khi chụp (ví dụ: 'EX1' hoặc 'EX')"""
        addr = (str(address).strip() if address else '') or 'EX1'
        if addr.isalpha():
            addr = f"{addr}1"
        self.goto_address = addr
    
    def set_zoom_level(self, zoom):
        """Đặt mức zoom cho Excel"""
        self.zoom_level = zoom
    
    def set_highlight_settings(self, fill_color, opacity):
        """Đặt màu và độ mờ cho highlight"""
        self.highlight_fill_color = fill_color
        self.highlight_fill_opacity = opacity
    
    @staticmethod
    def check_libraries():
        """Kiểm tra các thư viện cần thiết"""
        return {
            'pywin32': PYWIN32_AVAILABLE,
            'pywinauto': PYWINAUTO_AVAILABLE,
            'pil': PIL_AVAILABLE,
            'openpyxl': OPENPYXL_AVAILABLE,
            'can_run': PYWIN32_AVAILABLE and PYWINAUTO_AVAILABLE and PIL_AVAILABLE
        }
    
    def close_all_excel_instances(self):
        """Đóng tất cả Excel instances (legacy behavior)"""
        try:
            os.system("taskkill /f /im excel.exe")
            time.sleep(2)
            return True
        except Exception as e:
            utils.logger.error(f"Không thể đóng Excel: {e}")
            return False
    
    def _make_unique_sheet_name(self, workbook, desired_name):
        """Tạo tên sheet duy nhất để tránh trùng lặp"""
        base_name = desired_name
        counter = 1
        existing_names = set(sheet.Name for sheet in workbook.Sheets)
        
        while desired_name in existing_names:
            desired_name = f"{base_name}_{counter}"
            counter += 1
        
        return desired_name
    
    def open_excel_file(self, file_path):
        """Mở file Excel với COM Dispatch (Visible=True để chụp màn hình)"""
        try:
            pythoncom.CoInitialize()
            try:
                excel = win32com.client.Dispatch("Excel.Application")
            except Exception:
                excel = win32com.client.DispatchEx("Excel.Application")
            try:
                excel.Visible = True
            except Exception:
                pass
            try:
                excel.DisplayAlerts = False
            except Exception:
                pass
            try:
                excel.EnableEvents = False
            except Exception:
                pass
            try:
                excel.Application.AutomationSecurity = 3  # Disable macros
            except Exception:
                pass

            file_path = os.path.abspath(file_path)
            utils.unblock_file(file_path)
            workbook = excel.Workbooks.Open(
                file_path,
                UpdateLinks=False,
                ReadOnly=False,
                Notify=False,
                AddToMru=False
            )
            return excel, workbook
        except Exception as e:
            utils.logger.error(f"Không thể mở File Excel: {e}")
            return None, None
    
    def connect_to_excel_window(self, excel):
        """Kết nối với cửa sổ Excel qua PyWinAuto UIA"""
        if not PYWINAUTO_AVAILABLE:
            utils.logger.error("pywinauto không khả dụng")
            return None
        try:
            app = Application(backend='uia').connect(path=excel.Path, found_index=0)
            for w in app.windows():
                if "Excel" in w.window_text():
                    w.set_focus()
                    return app
            raise Exception("Không thể tìm thấy cửa sổ Excel")
        except Exception as e:
            utils.logger.error(f"Không thể kết nối với cửa sổ Excel: {e}")
            return None
    
    def close_excel(self, excel):
        """Đóng Excel và giải phóng tài nguyên"""
        try:
            if excel:
                excel.Quit()
        except Exception as e:
            utils.logger.debug(f"Lỗi khi đóng Excel: {e}")
    
    def check_and_turn_off_gridlines(self, excel):
        """Tắt gridlines trong Excel"""
        try:
            excel.ActiveWindow.DisplayGridlines = False
        except:
            pass
    
    def preprocess_excel_file(self, file_path, sheets_to_process=None, is_new=False, 
                               reference_name=None, status_callback=None):
        """
        Tiền xử lý file Excel theo logic Legacy:
        - Ẩn cột BE..EV
        - Lưu/áp dụng độ rộng cột
        - Đổi tên sheet nếu có dấu gạch dưới
        - Di chuyển tới ô chỉ định (goto_address)
        """
        excel = None
        try:
            if status_callback:
                status_callback(f"[Legacy] Đang tiền xử lý: {os.path.basename(file_path)}")
            
            try:
                excel = win32com.client.Dispatch("Excel.Application")
            except Exception:
                excel = win32com.client.DispatchEx("Excel.Application")
            try:
                excel.Visible = False
            except Exception:
                pass
            try:
                excel.DisplayAlerts = False
            except Exception:
                pass
            try:
                excel.EnableEvents = False
            except Exception:
                pass
            
            workbook = excel.Workbooks.Open(file_path, UpdateLinks=False, 
                                            ReadOnly=False, Notify=False, AddToMru=False)
            initial_sheet = workbook.ActiveSheet

            # Xác định sheets cần xử lý
            if sheets_to_process is not None:
                sheets_to_process_set = set([s.rstrip() for s in sheets_to_process])
                sheets = [sheet for sheet in workbook.Sheets 
                         if sheet.Name.rstrip() in sheets_to_process_set]
            else:
                sheets = [sheet for sheet in workbook.Sheets 
                         if sheet.Tab.Color == config.COLOR_GREEN_TAB]

            # Đổi tên sheet nếu có dấu gạch dưới
            for sheet in sheets:
                if "_" in sheet.Name:
                    desired = sheet.Name.replace("_", "-")
                    unique_name = self._make_unique_sheet_name(workbook, desired)
                    try:
                        sheet.Name = unique_name
                    except Exception:
                        pass

            total_sheets = len(sheets)

            # Nếu là file mới: lưu độ rộng cột A..BE (1..57)
            if is_new and total_sheets > 0:
                first_sheet = sheets[0]
                col_widths = []
                for col_idx in range(1, 58):  # 1..57 tương ứng A..BE
                    try:
                        cw = first_sheet.Columns(col_idx).ColumnWidth
                    except Exception:
                        cw = None
                    col_widths.append(cw)
                key = os.path.splitext(os.path.basename(file_path))[0]
                self.saved_col_widths[key] = col_widths

            for i, sheet in enumerate(sheets, 1):
                if status_callback:
                    status_callback(f"[Legacy] Xử lý sheet {i}/{total_sheets}")
                
                sheet.Activate()
                try:
                    excel.ActiveWindow.Zoom = self.zoom_level
                except Exception:
                    pass
                self.check_and_turn_off_gridlines(excel)

                # Ẩn cột BE..EV (57..152)
                try:
                    for col_idx in range(57, 153):
                        try:
                            sheet.Columns(col_idx).EntireColumn.Hidden = True
                        except Exception:
                            pass
                except Exception:
                    pass

                # Bỏ chọn vùng
                try:
                    excel.CutCopyMode = False
                except Exception:
                    pass
                    
                # Di chuyển con trỏ và cuộn màn hình đến ô chỉ định (goto_address, ví dụ EX1)
                goto_addr = (self.goto_address or 'EX1').strip()
                if goto_addr.isalpha():
                    goto_addr = f"{goto_addr}1"
                try:
                    excel.Application.Goto(Reference=sheet.Range(goto_addr), Scroll=True)
                except Exception:
                    try:
                        sheet.Range(goto_addr).Select()
                    except Exception:
                        pass

            # Nếu là file cũ: áp dụng độ rộng cột từ file mới
            if (not is_new) and reference_name:
                ref_key = os.path.splitext(os.path.basename(reference_name))[0]
                widths = self.saved_col_widths.get(ref_key)
                if widths:
                    for sheet in sheets:
                        for col_idx, w in enumerate(widths, start=1):
                            try:
                                if w is not None:
                                    sheet.Columns(col_idx).ColumnWidth = w
                            except Exception:
                                pass

            initial_sheet.Activate()
            workbook.Save()
            workbook.Close(True)
            excel.Quit()
            
            if status_callback:
                status_callback(f"[Legacy] Đã tiền xử lý xong: {os.path.basename(file_path)}")
            return True
            
        except Exception as e:
            utils.logger.error(f"Lỗi tiền xử lý {os.path.basename(file_path)}: {e}")
            return False
        finally:
            try:
                if excel:
                    excel.Quit()
            except:
                pass
    
    def process_sheet(self, excel, app, sheet, output_prefix, base_path):
        """
        Chụp ảnh một sheet theo logic Legacy.
        Sử dụng screen mode để xác định vùng cắt.
        """
        utils.logger.info(f"[Legacy] Processing sheet: {sheet.Name}")
        sheet.Activate()
        
        # Di chuyển con trỏ và cuộn màn hình đến ô chỉ định (goto_address, ví dụ EX1)
        goto_addr = (self.goto_address or 'EX1').strip()
        if goto_addr.isalpha():
            goto_addr = f"{goto_addr}1"
            
        try:
            excel.Application.Goto(Reference=sheet.Range(goto_addr), Scroll=True)
            utils.logger.info(f"[Legacy] Moved cursor & scrolled to {goto_addr}")
        except Exception as goto_err:
            try:
                sheet.Range(goto_addr).Select()
                excel.ActiveWindow.ScrollRow = 1
                excel.ActiveWindow.ScrollColumn = sheet.Range(goto_addr).Column
            except Exception:
                pass
                
        try:
            excel.ActiveWindow.Zoom = self.zoom_level
        except Exception:
            pass
        
        try:
            excel.ExecuteExcel4Macro("SHOW.TOOLBAR(\"Ribbon\",False)")
        except:
            pass
        
        self.check_and_turn_off_gridlines(excel)
        time.sleep(1)
        
        bbox = pyautogui.screenshot().getbbox()
        if bbox:
            selected_width = bbox[2] - bbox[0]
            selected_height = bbox[3] - bbox[1]
            
            mode = self.screen_mode
            if mode == 'pc':
                start_x = bbox[0] + int(selected_width / 3)
                end_x = bbox[0] + int(selected_width * 2 / 3)
            elif mode == 'vps':
                start_x = bbox[0] + int(selected_width * 2 / 6)
                end_x = bbox[0] + int(selected_width * 5 / 6)
            elif mode == 'monitor':
                start_x = bbox[0] + int(selected_width / 5)
                end_x = bbox[0] + int(selected_width * 4 / 5)
            else:
                start_x = bbox[0] + int(selected_width / 3)
                end_x = bbox[0] + int(selected_width * 2 / 3)
            
            new_bbox = (start_x, bbox[1], end_x, bbox[3])
            img = ImageGrab.grab(bbox=new_bbox)
            
            sheet_name_stripped = sheet.Name.rstrip()
            output_path = os.path.join(base_path, f"{output_prefix}_{sheet_name_stripped}.png")
            img.save(output_path)
            utils.logger.info(f"[Legacy] Screenshot saved: {output_path}")
            return output_path
        return None
    
    def process_excel_file(self, file_path, output_prefix, base_path, 
                           sheets_to_process=None, status_callback=None):
        """
        Xử lý file Excel: mở, chụp ảnh các sheet, đóng.
        """
        excel, workbook = self.open_excel_file(file_path)
        if not excel or not workbook:
            return False

        app = self.connect_to_excel_window(excel)
        if not app:
            self.close_excel(excel)
            return False

        processed_sheets = []

        try:
            workbook.Activate()
            excel.Application.WindowState = -4137  # xlMaximized
            
            try:
                excel.ExecuteExcel4Macro("SHOW.TOOLBAR(\"Ribbon\",True)")
            except:
                pass

            time.sleep(3)

            for sheet in workbook.Sheets:
                sheet_name_stripped = sheet.Name.rstrip()
                
                if output_prefix == "CTTTmoi":
                    if sheet.Tab.Color == config.COLOR_GREEN_TAB:
                        # Đổi tên sheet nếu có dấu gạch dưới
                        if "_" in sheet_name_stripped:
                            try:
                                new_name = sheet_name_stripped.replace("_", "-")
                                new_name = self._make_unique_sheet_name(workbook, new_name)
                                sheet.Name = new_name
                                sheet_name_stripped = new_name
                            except Exception:
                                pass
                        
                        processed_sheets.append(sheet_name_stripped)
                        self.process_sheet(excel, app, sheet, output_prefix, base_path)
                        
                elif sheets_to_process and sheet_name_stripped in [s.rstrip() for s in sheets_to_process]:
                    self.process_sheet(excel, app, sheet, output_prefix, base_path)

        except Exception as e:
            utils.logger.error(f"Lỗi xử lý file Excel: {e}")
            return False

        finally:
            workbook.Close(SaveChanges=False)
            self.close_excel(excel)
        
        return processed_sheets if output_prefix == "CTTTmoi" else True
    
    def hex_to_rgba(self, hex_color, opacity):
        """Chuyển đổi màu hex sang RGBA"""
        hex_color = hex_color.lstrip('#')
        lv = len(hex_color)
        r, g, b = tuple(int(hex_color[i:i + lv // 3], 16) for i in range(0, lv, lv // 3))
        a = int(255 * (opacity / 100))
        return (r, g, b, a)
    
    def compare_images(self, new_file, base_path):
        """
        So sánh các ảnh đã chụp (Legacy method - pixel by pixel).
        """
        excel_file1_name = os.path.splitext(os.path.basename(new_file))[0]

        new_images = [f for f in os.listdir(base_path) 
                     if f.startswith("CTTTmoi_") and f.endswith(".png")]
        old_images = [f for f in os.listdir(base_path) 
                     if f.startswith("CTTTcu_") and f.endswith(".png")]

        utils.logger.info(f"[Legacy] New images: {len(new_images)}, Old images: {len(old_images)}")

        # Map tên sheet đã rstrip tới file ảnh
        new_img_map = {f.split("_", 1)[1].rsplit(".png", 1)[0].rstrip(): f for f in new_images}
        old_img_map = {f.split("_", 1)[1].rsplit(".png", 1)[0].rstrip(): f for f in old_images}

        comparison_results = {}

        for sheet_name in new_img_map:
            if sheet_name in old_img_map:
                new_img = new_img_map[sheet_name]
                old_img = old_img_map[sheet_name]
                img1_path = os.path.join(base_path, new_img)
                img2_path = os.path.join(base_path, old_img)

                try:
                    img1 = Image.open(img1_path)
                    img2 = Image.open(img2_path)
                    img2 = img2.resize(img1.size)
                    
                    diff = ImageChops.difference(img1, img2)
                    
                    if diff.getbbox():
                        utils.logger.info(f"[Legacy] Differences detected: {sheet_name}")
                        
                        highlight = Image.new("RGBA", img1.size, (255, 255, 255, 0))
                        mask = diff.convert("L").point(lambda x: min(x, 500))
                        draw = ImageDraw.Draw(highlight)
                        
                        highlight_fill_rgba = self.hex_to_rgba(
                            self.highlight_fill_color, 
                            self.highlight_fill_opacity
                        )
                        
                        # Vẽ pixel by pixel (legacy behavior)
                        for x in range(img1.width):
                            for y in range(img1.height):
                                if mask.getpixel((x, y)) > 0:
                                    draw.rectangle([x, y, x + 1, y + 1], fill=highlight_fill_rgba)
                        
                        combined = Image.alpha_composite(img1.convert("RGBA"), highlight)
                        result_filename = f"{excel_file1_name}_{sheet_name}_so_sanh.png"
                        result_path = os.path.join(base_path, result_filename)
                        combined.save(result_path)
                        
                        comparison_results[sheet_name] = result_path
                        utils.logger.info(f"[Legacy] Comparison saved: {result_path}")
                    else:
                        utils.logger.info(f"[Legacy] No differences: {sheet_name}")
                        comparison_results[sheet_name] = None
                        
                except Exception as e:
                    utils.logger.error(f"[Legacy] Error comparing {sheet_name}: {e}")
            else:
                utils.logger.warning(f"[Legacy] No old image for: {sheet_name}")

        return comparison_results
    
    def create_excel_result(self, new_file, base_path, output_folder=None):
        """
        Tạo file Excel kết quả với ảnh so sánh (Legacy format).
        """
        if not OPENPYXL_AVAILABLE:
            utils.logger.error("openpyxl không khả dụng")
            return None
            
        excel_filename = os.path.basename(new_file)
        excel_name_only, excel_ext = os.path.splitext(excel_filename)
        excel_result_filename = f"Kết quả_{excel_name_only}.xlsx"

        if output_folder and os.path.isdir(output_folder):
            excel_result_path = os.path.join(output_folder, excel_result_filename)
        else:
            excel_result_path = os.path.join(base_path, excel_result_filename)

        comparison_images = [f for f in os.listdir(base_path) 
                           if f.startswith(excel_name_only) and f.endswith("_so_sanh.png")]
        cu_images = [f for f in os.listdir(base_path) 
                    if f.startswith("CTTTcu_") and f.endswith(".png")]

        if not comparison_images:
            utils.logger.warning(f"Không tìm thấy ảnh so sánh cho {excel_filename}")
            return None

        wb = Workbook()
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        for img_file in comparison_images:
            sheet_name = img_file.split("_so_sanh")[0].split("_", 1)[1]
            ws = wb.create_sheet(title=f" {sheet_name}"[:31])
            
            # Add CTTTcu image
            cu_img_file = f"CTTTcu_{sheet_name}.png"
            if cu_img_file in cu_images:
                cu_img_path = os.path.join(base_path, cu_img_file)
                try:
                    cu_img = xlImage(cu_img_path)
                    ws.add_image(cu_img, 'A1')
                except Exception as e:
                    utils.logger.debug(f"Cannot add old image: {e}")
            
            # Add comparison image
            comparison_img_path = os.path.join(base_path, img_file)
            try:
                comparison_img = xlImage(comparison_img_path)
                ws.add_image(comparison_img, 'G1')
            except Exception as e:
                utils.logger.debug(f"Cannot add comparison image: {e}")

        wb.save(excel_result_path)
        utils.logger.info(f"[Legacy] Excel result saved: {excel_result_path}")
        
        return excel_result_path
    
    def cleanup_temp_images(self, base_path):
        """Xóa các file ảnh tạm"""
        patterns = ["CTTTmoi_*.png", "CTTTcu_*.png", "*_so_sanh.png"]
        for pattern in patterns:
            for f in glob.glob(os.path.join(base_path, pattern)):
                try:
                    os.remove(f)
                except:
                    pass
    
    def run_full_comparison(self, new_files, old_files, output_folder=None, 
                            status_callback=None, progress_callback=None):
        """
        Chạy toàn bộ quy trình so sánh Legacy.
        
        Args:
            new_files: List đường dẫn file CTTT mới
            old_files: List đường dẫn file CTTT cũ
            output_folder: Thư mục lưu kết quả (None = cùng thư mục với file mới)
            status_callback: Hàm callback cập nhật trạng thái
            progress_callback: Hàm callback cập nhật tiến độ (0-100)
            
        Returns:
            Tuple (success: bool, elapsed_time: float, results: list)
        """
        if len(new_files) != len(old_files):
            utils.logger.error("Số lượng file mới và cũ không khớp")
            return False, 0, []
        
        start_time = time.time()
        total_files = len(new_files)
        all_results = []
        
        for index, (new_file, old_file) in enumerate(zip(new_files, old_files), 1):
            if status_callback:
                status_callback(f"[Legacy] Đang xử lý cặp {index}/{total_files}")
            
            base_path = output_folder or os.path.dirname(new_file)
            
            try:
                # Lấy danh sách sheet có màu xanh từ file mới
                try:
                    excel = win32com.client.Dispatch("Excel.Application")
                except Exception:
                    excel = win32com.client.DispatchEx("Excel.Application")
                try:
                    excel.Visible = False
                except Exception:
                    pass
                try:
                    excel.DisplayAlerts = False
                except Exception:
                    pass
                
                sheet_names_colored = []
                try:
                    wb = excel.Workbooks.Open(new_file, UpdateLinks=False, 
                                              ReadOnly=True, Notify=False, AddToMru=False)
                    for sheet in wb.Sheets:
                        try:
                            if int(sheet.Tab.Color) == config.COLOR_GREEN_TAB:
                                sheet_names_colored.append(sheet.Name.rstrip())
                        except Exception:
                            continue
                    wb.Close(False)
                except Exception as open_err:
                    utils.logger.error(f"Lỗi khi đọc sheet từ {new_file}: {open_err}")
                finally:
                    try:
                        excel.Quit()
                    except Exception:
                        pass
                
                if not sheet_names_colored:
                    utils.logger.warning(f"Không tìm thấy sheet màu xanh: {new_file}")
                    continue
                
                # Tiền xử lý file mới
                if status_callback:
                    status_callback(f"[Legacy] Tiền xử lý file mới...")
                if not self.preprocess_excel_file(new_file, sheets_to_process=sheet_names_colored,
                                                   is_new=True, status_callback=status_callback):
                    continue
                
                # Tiền xử lý file cũ
                if status_callback:
                    status_callback(f"[Legacy] Tiền xử lý file cũ...")
                if not self.preprocess_excel_file(old_file, sheets_to_process=sheet_names_colored,
                                                   is_new=False, reference_name=new_file,
                                                   status_callback=status_callback):
                    continue
                
                # Chụp ảnh file mới
                if status_callback:
                    status_callback(f"[Legacy] Chụp ảnh file mới: {os.path.basename(new_file)}")
                processed_sheets = self.process_excel_file(new_file, "CTTTmoi", base_path)
                if not processed_sheets:
                    continue
                
                time.sleep(2)
                
                # Chụp ảnh file cũ
                if status_callback:
                    status_callback(f"[Legacy] Chụp ảnh file cũ: {os.path.basename(old_file)}")
                success_old = self.process_excel_file(old_file, "CTTTcu", base_path, processed_sheets)
                if not success_old:
                    continue
                
                # So sánh ảnh
                if status_callback:
                    status_callback("[Legacy] Đang so sánh ảnh...")
                comparison_results = self.compare_images(new_file, base_path)
                
                # Tạo file Excel kết quả
                if status_callback:
                    status_callback("[Legacy] Đang tạo file Excel kết quả...")
                result_path = self.create_excel_result(new_file, base_path, output_folder)
                
                if result_path:
                    all_results.append(result_path)
                
                # Dọn dẹp ảnh tạm
                self.cleanup_temp_images(base_path)
                
            except Exception as e:
                utils.logger.error(f"[Legacy] Error processing {new_file}: {e}")
            
            # Update progress
            if progress_callback:
                progress = (index / total_files) * 100
                progress_callback(progress)
        
        elapsed_time = time.time() - start_time
        
        if status_callback:
            minutes, seconds = divmod(elapsed_time, 60)
            status_callback(f"[Legacy] Hoàn tất! Thời gian: {int(minutes)} phút {seconds:.2f} giây")
        
        return True, elapsed_time, all_results
