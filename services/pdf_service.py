import os
import time
import shutil
import uuid
import win32com.client
import pythoncom
import utils
import config

# Library availability flags
PYMUPDF_AVAILABLE = False
PYPDF2_AVAILABLE = False
REPORTLAB_AVAILABLE = False
PDF2IMAGE_AVAILABLE = False

try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    pass

try:
    import PyPDF2
    from PyPDF2 import PdfMerger, PdfReader, PdfWriter
    PYPDF2_AVAILABLE = True
except ImportError:
    pass

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.utils import ImageReader
    from reportlab.lib.pagesizes import A4, landscape
    REPORTLAB_AVAILABLE = True
except ImportError:
    pass

try:
    from pdf2image import convert_from_path
    PDF2IMAGE_AVAILABLE = True
except ImportError:
    convert_from_path = None

from PIL import Image, ImageChops, ImageFilter

# Import optimized image comparison module
try:
    from services.optimized_image_compare import (
        compare_images_opencv, create_side_by_side, quick_compare_hash, is_opencv_available
    )
    OPENCV_COMPARE_AVAILABLE = True
except ImportError:
    OPENCV_COMPARE_AVAILABLE = False


class PDFService:
    """
    Comprehensive PDF Service for:
    - Exporting Excel sheets to PDF
    - Merging multiple PDFs
    - Rendering PDF pages to images (PyMuPDF)
    - Comparing PDF pages and creating highlight overlays
    - Creating side-by-side comparison PDFs
    """
    
    def __init__(self):
        self.excel_app = None
        self.excel_pid = None
    
    # =========================================================================
    # LIBRARY CHECKS
    # =========================================================================
    
    @staticmethod
    def check_pdf_libraries():
        """
        Check which PDF libraries are available.
        Returns dict with availability status and error messages.
        """
        status = {
            'pymupdf': PYMUPDF_AVAILABLE,
            'pypdf2': PYPDF2_AVAILABLE,
            'reportlab': REPORTLAB_AVAILABLE,
            'pdf2image': PDF2IMAGE_AVAILABLE,
        }
        
        missing = []
        if not PYMUPDF_AVAILABLE:
            missing.append("PyMuPDF (pip install pymupdf)")
        if not REPORTLAB_AVAILABLE:
            missing.append("reportlab (pip install reportlab)")
            
        status['missing_critical'] = missing
        status['can_compare'] = PYMUPDF_AVAILABLE and REPORTLAB_AVAILABLE
        
        return status
    
    # =========================================================================
    # EXCEL TO PDF EXPORT
    # =========================================================================
    
    def _init_excel(self):
        """Initialize Excel COM application with performance optimizations."""
        if self.excel_app is not None:
            try:
                _ = self.excel_app.Workbooks.Count
            except Exception:
                utils.logger.info("Existing Excel COM instance stale, reinitializing...")
                self._cleanup_excel()

        if self.excel_app is None:
            pythoncom.CoInitialize()
            try:
                self.excel_app = win32com.client.DispatchEx("Excel.Application")
            except Exception as e:
                utils.logger.warning(f"DispatchEx failed, trying Dispatch: {e}")
                try:
                    self.excel_app = win32com.client.Dispatch("Excel.Application")
                except Exception as e2:
                    utils.logger.error(f"Failed to create Excel.Application instance: {e2}")
                    raise
            
            try:
                self.excel_app.Visible = False
            except Exception as e:
                utils.logger.warning(f"Could not set Visible=False on Excel.Application: {e}")
                
            try:
                self.excel_app.DisplayAlerts = False
            except Exception:
                pass
                
            try:
                self.excel_app.EnableEvents = False
            except Exception:
                pass
                
            try:
                self.excel_app.Interactive = False
            except Exception:
                pass
            
            # === PERFORMANCE OPTIMIZATION (Cấp độ 1) ===
            # Tắt cập nhật màn hình và tính toán tự động
            # Bọc trong try-except vì một số phiên bản Excel có thể từ chối
            try:
                self.excel_app.ScreenUpdating = False
            except Exception:
                pass
            
            try:
                self.excel_app.Calculation = -4135  # xlCalculationManual
            except Exception:
                pass
            
            try:
                self.excel_app.AskToUpdateLinks = False
            except Exception:
                pass
            
            try:
                self.excel_pid = utils.get_excel_pid(self.excel_app)
            except Exception:
                self.excel_pid = None

    
    def _cleanup_excel(self):
        """Cleanup Excel COM application."""
        if self.excel_app:
            try:
                self.excel_app.Quit()
            except:
                pass
            self.excel_app = None
        if self.excel_pid:
            utils.kill_specific_excel_process(self.excel_pid)
            self.excel_pid = None
    
    def get_colored_sheets(self, file_path):
        """
        Get list of sheet names with green tab color (5296274).
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            List of sheet names with correct tab color
        """
        colored_sheets = []
        workbook = None
        try:
            self._init_excel()
            
            # Unblock file to avoid Protected View issues
            utils.unblock_file(file_path)
            
            workbook = self.excel_app.Workbooks.Open(
                file_path, 
                ReadOnly=True, 
                UpdateLinks=False,
                Notify=False,
                AddToMru=False
            )
            
            for sheet in workbook.Sheets:
                try:
                    if int(sheet.Tab.Color) == config.COLOR_GREEN_TAB:
                        colored_sheets.append(sheet.Name.rstrip())
                except Exception:
                    continue
            
            utils.logger.info(f"Found {len(colored_sheets)} green-tab sheets: {colored_sheets}")
            
        except Exception as e:
            utils.logger.error(f"Error getting colored sheets: {e}")
        finally:
            if workbook:
                try:
                    workbook.Close(SaveChanges=False)
                except Exception:
                    pass
        
        return colored_sheets

    def get_visible_sheets(self, file_path):
        """
        Lấy danh sách các Sheet đang hiển thị (Visible != 0 / xlSheetHidden).
        Tự động bỏ qua các sheet ẩn hệ thống và sheet trống hoàn toàn.
        """
        import time
        visible_sheets = []
        for attempt in range(2):
            workbook = None
            try:
                self._init_excel()
                utils.unblock_file(file_path)
                
                workbook = self.excel_app.Workbooks.Open(
                    file_path, 
                    ReadOnly=True, 
                    UpdateLinks=False, 
                    Notify=False, 
                    AddToMru=False
                )
                
                for sheet in workbook.Sheets:
                    try:
                        is_visible = (sheet.Visible == -1 or sheet.Visible is True or sheet.Visible == 1)
                        if is_visible:
                            s_name = sheet.Name.rstrip()
                            try:
                                used = sheet.UsedRange
                                if used.Rows.Count == 1 and used.Columns.Count == 1:
                                    if used.Value is None or str(used.Value).strip() == "":
                                        utils.logger.info(f"Skipping empty sheet: '{s_name}'")
                                        continue
                            except Exception:
                                pass
                            visible_sheets.append(s_name)
                    except Exception as e:
                        utils.logger.warning(f"Error checking sheet visibility: {e}")
                        continue
                        
                utils.logger.info(f"Found {len(visible_sheets)} visible sheets in {os.path.basename(file_path)}: {visible_sheets}")
                break
            except Exception as e:
                utils.logger.error(f"Error getting visible sheets from {file_path} (attempt {attempt+1}): {e}")
                self._cleanup_excel()
                if attempt == 0:
                    time.sleep(1)
            finally:
                if workbook:
                    try:
                        workbook.Close(SaveChanges=False)
                    except Exception:
                        pass
        return visible_sheets

    def get_visible_sheets_fast(self, file_path):
        """Read visible sheet names from XLSX/XLSM metadata without opening Excel."""
        try:
            from services.validation_service import ValidationService

            metadata = ValidationService._inspect_workbook(file_path)
            if metadata is None:
                return None
            visible = [
                item["name"].rstrip()
                for item in metadata
                if item["visible"] and item.get("has_content", True)
            ]
            utils.logger.info(
                f"Found {len(visible)} visible sheets from workbook metadata: {visible}"
            )
            return visible
        except Exception as error:
            utils.logger.warning(f"Fast visible-sheet discovery failed for {file_path}: {error}")
            return None
    
    def find_sheet_by_name(self, file_path, target_name="form"):
        """
        Tìm chính xác tên của Sheet theo tên target (không phân biệt hoa thường và khoảng trắng).
        Trả về tên gốc của Sheet trong Excel nếu tìm thấy, ngược lại trả về None.
        """
        found_name = None
        for attempt in range(2):
            workbook = None
            try:
                self._init_excel()
                utils.unblock_file(file_path)
                workbook = self.excel_app.Workbooks.Open(
                    file_path,
                    ReadOnly=True,
                    UpdateLinks=False,
                    Notify=False,
                    AddToMru=False
                )
                target_clean = target_name.strip().lower()
                for sheet in workbook.Sheets:
                    try:
                        s_name = sheet.Name.rstrip()
                        if s_name.strip().lower() == target_clean:
                            found_name = s_name
                            break
                    except Exception:
                        continue
                if found_name:
                    break
            except Exception as e:
                utils.logger.error(f"Error finding sheet '{target_name}' in {file_path}: {e}")
                self._cleanup_excel()
                if attempt == 0:
                    time.sleep(1)
            finally:
                if workbook:
                    try:
                        workbook.Close(SaveChanges=False)
                    except Exception:
                        pass
        return found_name

    @staticmethod
    def _find_sheet(workbook, sheet_name):
        """Find a worksheet while tolerating trailing spaces and case differences."""
        target = sheet_name.strip().lower()
        for sheet in workbook.Sheets:
            if sheet.Name.strip().lower() == target:
                return sheet
        return None

    @staticmethod
    def _sync_sheet_layout(target_sheet, reference_sheet, print_area):
        """Normalize row/column geometry in memory before a visual comparison.

        Excel stores widths and heights as fractional values. Tiny differences between
        workbook revisions reflow every text line in the exported PDF and otherwise
        create thousands of false-positive pixels.
        """
        reference_range = reference_sheet.Range(print_area)
        target_range = target_sheet.Range(print_area)

        for index in range(1, reference_range.Columns.Count + 1):
            source = reference_range.Columns(index).EntireColumn
            destination = target_range.Columns(index).EntireColumn
            destination.ColumnWidth = source.ColumnWidth
            destination.Hidden = source.Hidden

        for index in range(1, reference_range.Rows.Count + 1):
            source = reference_range.Rows(index).EntireRow
            destination = target_range.Rows(index).EntireRow
            destination.RowHeight = source.RowHeight
            destination.Hidden = source.Hidden

    @staticmethod
    def _read_sheet_layout(file_path, sheet_name, print_area):
        """Read effective row/column geometry without starting or querying Excel COM."""
        import warnings
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter, range_boundaries

        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Data Validation extension is not supported",
                category=UserWarning,
            )
            workbook = load_workbook(
                file_path,
                read_only=False,
                data_only=False,
                keep_links=False,
            )
        try:
            target = sheet_name.strip().lower()
            worksheet = next(
                (sheet for sheet in workbook.worksheets if sheet.title.strip().lower() == target),
                None,
            )
            if worksheet is None:
                return None

            min_col, min_row, max_col, max_row = range_boundaries(print_area)
            default_width = worksheet.sheet_format.defaultColWidth
            if default_width is None:
                default_width = worksheet.sheet_format.baseColWidth or 8.43
            # Excel ignores defaultRowHeight unless customHeight is explicitly set.
            default_height = (
                worksheet.sheet_format.defaultRowHeight
                if worksheet.sheet_format.customHeight
                else 15.0
            )

            column_dimensions = list(worksheet.column_dimensions.values())

            columns = []
            for column_index in range(min_col, max_col + 1):
                dimension = next(
                    (
                        item for item in column_dimensions
                        if item.min <= column_index <= item.max
                    ),
                    None,
                )
                columns.append({
                    "index": column_index,
                    "size": dimension.width if dimension and dimension.width is not None else default_width,
                    "hidden": bool(dimension.hidden) if dimension else False,
                })

            rows = []
            for row_index in range(min_row, max_row + 1):
                dimension = worksheet.row_dimensions.get(row_index)
                rows.append({
                    "index": row_index,
                    "size": dimension.height if dimension and dimension.height is not None else default_height,
                    "hidden": bool(dimension.hidden) if dimension else False,
                })

            return {"columns": columns, "rows": rows}
        finally:
            workbook.close()

    @staticmethod
    def _group_layout_dimensions(dimensions):
        """Group adjacent dimensions with identical geometry for batched COM writes."""
        groups = []
        for dimension in dimensions:
            key = (round(float(dimension["size"]), 8), bool(dimension["hidden"]))
            if groups and groups[-1]["end"] + 1 == dimension["index"] and groups[-1]["key"] == key:
                groups[-1]["end"] = dimension["index"]
            else:
                groups.append({
                    "start": dimension["index"],
                    "end": dimension["index"],
                    "key": key,
                })
        return groups

    @classmethod
    def _apply_sheet_layout(cls, target_sheet, layout, current_layout=None,
                            apply_columns=True, apply_rows=True):
        """Apply pre-read layout using one COM operation per contiguous size group."""
        from openpyxl.utils import get_column_letter

        def changed(reference, current, field):
            if current is None:
                return reference
            current_by_index = {item["index"]: item for item in current}
            result = []
            for item in reference:
                old = current_by_index.get(item["index"])
                if old is None:
                    result.append(item)
                elif field == "size" and abs(float(item["size"]) - float(old["size"])) > 1e-8:
                    result.append(item)
                elif field == "hidden" and bool(item["hidden"]) != bool(old["hidden"]):
                    result.append(item)
            return result

        if apply_columns:
            column_sizes = changed(
                layout["columns"], current_layout["columns"] if current_layout else None, "size"
            )
            for group in cls._group_layout_dimensions(column_sizes):
                start = get_column_letter(group["start"])
                end = get_column_letter(group["end"])
                target = target_sheet.Range(f"{start}:{end}").EntireColumn
                target.ColumnWidth = group["key"][0]

            column_visibility = changed(
                layout["columns"], current_layout["columns"] if current_layout else None, "hidden"
            )
            for group in cls._group_layout_dimensions(column_visibility):
                start = get_column_letter(group["start"])
                end = get_column_letter(group["end"])
                target_sheet.Range(f"{start}:{end}").EntireColumn.Hidden = group["key"][1]

        if apply_rows:
            row_sizes = changed(
                layout["rows"], current_layout["rows"] if current_layout else None, "size"
            )
            for group in cls._group_layout_dimensions(row_sizes):
                target = target_sheet.Range(f"{group['start']}:{group['end']}").EntireRow
                target.RowHeight = group["key"][0]

            row_visibility = changed(
                layout["rows"], current_layout["rows"] if current_layout else None, "hidden"
            )
            for group in cls._group_layout_dimensions(row_visibility):
                target_sheet.Range(f"{group['start']}:{group['end']}").EntireRow.Hidden = group["key"][1]

    @classmethod
    def _sync_changed_layout_from_com(cls, target_sheet, reference_sheet,
                                      reference_layout, current_layout, print_area):
        """Sync only changed geometry groups, using Excel's effective display units."""
        from openpyxl.utils import get_column_letter, range_boundaries

        min_col, min_row, _, _ = range_boundaries(print_area)
        reference_range = reference_sheet.Range(print_area)

        def changed(reference, current, field):
            current_by_index = {item["index"]: item for item in current}
            result = []
            for item in reference:
                old = current_by_index.get(item["index"])
                if old is None:
                    result.append(item)
                elif field == "size" and abs(float(item["size"]) - float(old["size"])) > 1e-8:
                    result.append(item)
                elif field == "hidden" and bool(item["hidden"]) != bool(old["hidden"]):
                    result.append(item)
            return result

        changed_columns = changed(reference_layout["columns"], current_layout["columns"], "size")
        for group in cls._group_layout_dimensions(changed_columns):
            relative_index = group["start"] - min_col + 1
            effective_width = reference_range.Columns(relative_index).EntireColumn.ColumnWidth
            start = get_column_letter(group["start"])
            end = get_column_letter(group["end"])
            target_sheet.Range(f"{start}:{end}").EntireColumn.ColumnWidth = effective_width

        hidden_columns = changed(reference_layout["columns"], current_layout["columns"], "hidden")
        for group in cls._group_layout_dimensions(hidden_columns):
            relative_index = group["start"] - min_col + 1
            hidden = reference_range.Columns(relative_index).EntireColumn.Hidden
            start = get_column_letter(group["start"])
            end = get_column_letter(group["end"])
            target_sheet.Range(f"{start}:{end}").EntireColumn.Hidden = hidden

        changed_rows = changed(reference_layout["rows"], current_layout["rows"], "size")
        for group in cls._group_layout_dimensions(changed_rows):
            relative_index = group["start"] - min_row + 1
            effective_height = reference_range.Rows(relative_index).EntireRow.RowHeight
            target_sheet.Range(f"{group['start']}:{group['end']}").EntireRow.RowHeight = effective_height

        hidden_rows = changed(reference_layout["rows"], current_layout["rows"], "hidden")
        for group in cls._group_layout_dimensions(hidden_rows):
            relative_index = group["start"] - min_row + 1
            hidden = reference_range.Rows(relative_index).EntireRow.Hidden
            target_sheet.Range(f"{group['start']}:{group['end']}").EntireRow.Hidden = hidden

    def export_sheets_to_pdf(self, file_path, sheet_names, output_pdf_path, print_area="EX1:GR76",
                             _keep_alive=False, layout_reference_path=None):
        """
        Export specified sheets from Excel file to a single PDF.
        
        Args:
            file_path: Path to Excel file
            sheet_names: List of sheet names to export
            output_pdf_path: Output PDF path
            print_area: Print area range (default: "EX1:GR76")
            
        Returns:
            True if successful, False otherwise
        """
        temp_pdf_files = []
        exported_page_labels = []
        output_dir = os.path.dirname(output_pdf_path)
        workbook = None  # Track for cleanup in finally
        reference_workbook = None
        reference_layouts = {}
        current_layouts = {}
        
        try:
            self._init_excel()
            
            # Unblock file
            utils.unblock_file(file_path)
            
            workbook = self.excel_app.Workbooks.Open(
                file_path,
                ReadOnly=True,
                UpdateLinks=False,
                Notify=False,
                AddToMru=False
            )

            if layout_reference_path:
                for sheet_name in sheet_names:
                    try:
                        key = sheet_name.strip().lower()
                        reference_layouts[key] = self._read_sheet_layout(
                            layout_reference_path, sheet_name, print_area
                        )
                        current_layouts[key] = self._read_sheet_layout(
                            file_path, sheet_name, print_area
                        )
                    except Exception as layout_error:
                        utils.logger.warning(
                            f"Could not read reference layout for '{sheet_name}': {layout_error}"
                        )
                utils.unblock_file(layout_reference_path)
                reference_workbook = self.excel_app.Workbooks.Open(
                    layout_reference_path,
                    ReadOnly=True,
                    UpdateLinks=False,
                    Notify=False,
                    AddToMru=False,
                )
            
            # Find and export each sheet
            for idx, sheet_name in enumerate(sheet_names, 1):
                sheet = None
                
                # Try exact match first
                for s in workbook.Sheets:
                    if s.Name.rstrip() == sheet_name:
                        sheet = s
                        break
                
                # Fallback 1: Try variant names (underscore <-> dash)
                if sheet is None:
                    # Try replacing dash with underscore
                    variant1 = sheet_name.replace("-", "_")
                    # Try replacing underscore with dash
                    variant2 = sheet_name.replace("_", "-")
                    
                    for s in workbook.Sheets:
                        s_name = s.Name.rstrip()
                        if s_name == variant1 or s_name == variant2:
                            sheet = s
                            utils.logger.info(f"Found sheet with variant name: '{s_name}' for '{sheet_name}'")
                            break

                # Fallback 2: Try case-insensitive matching
                if sheet is None:
                    target_lower = sheet_name.strip().lower()
                    for s in workbook.Sheets:
                        if s.Name.strip().lower() == target_lower:
                            sheet = s
                            utils.logger.info(f"Found sheet with case-insensitive name: '{s.Name}' for '{sheet_name}'")
                            break
                
                if sheet is None:
                    utils.logger.warning(f"Sheet '{sheet_name}' not found in {file_path}")
                    continue

                reference_layout = reference_layouts.get(sheet_name.strip().lower())
                if reference_layout:
                    try:
                        reference_sheet = self._find_sheet(reference_workbook, sheet_name)
                        if reference_sheet:
                            self._sync_changed_layout_from_com(
                                sheet,
                                reference_sheet,
                                reference_layout,
                                current_layouts.get(sheet_name.strip().lower()),
                                print_area,
                            )
                        else:
                            self._apply_sheet_layout(
                                sheet,
                                reference_layout,
                                current_layouts.get(sheet_name.strip().lower()),
                            )
                        self.excel_app.CutCopyMode = False
                        utils.logger.info(
                            f"Synchronized layout for '{sheet_name}' from comparison reference"
                        )
                    except Exception as sync_error:
                        utils.logger.warning(
                            f"Could not synchronize layout for '{sheet_name}': {sync_error}"
                        )
                
                # Setup PageSetup for export (essential properties only for maximum speed)
                try:
                    sheet.Activate()
                    ps = sheet.PageSetup
                    
                    # Set print area & essential fit: Fit width & height to 1 page (do not split into multiple pages)
                    ps.PrintArea = print_area
                    ps.PaperSize = 9    # A4
                    ps.Zoom = False
                    ps.FitToPagesWide = 1
                    ps.FitToPagesTall = 1
                    ps.CenterHorizontally = True
                except Exception as e:
                    utils.logger.error(f"Error setting up sheet '{sheet_name}': {e}")
                
                # Export to temp PDF
                safe_name = utils.sanitize_filename_strict(sheet_name)
                unique_id = str(uuid.uuid4())[:8]
                temp_pdf = os.path.join(output_dir, f"temp_sheet_{idx}_{safe_name}_{unique_id}.pdf")
                
                try:
                    sheet.Select()
                    sheet.ExportAsFixedFormat(
                        Type=0,  # xlTypePDF
                        Filename=temp_pdf,
                        Quality=0,
                        IncludeDocProperties=True,
                        IgnorePrintAreas=False,
                        OpenAfterPublish=False
                    )
                    
                    if os.path.exists(temp_pdf):
                        temp_pdf_files.append(temp_pdf)
                        try:
                            import fitz
                            doc = fitz.open(temp_pdf)
                            p_cnt = len(doc)
                            doc.close()
                        except Exception:
                            p_cnt = 1
                        if p_cnt <= 1:
                            exported_page_labels.append(sheet_name)
                        else:
                            for p in range(p_cnt):
                                exported_page_labels.append(f"{sheet_name}_P{p+1}")
                    else:
                        utils.logger.error(f"Failed to create PDF for sheet '{sheet_name}'")
                        
                except Exception as e:
                    utils.logger.error(f"Error exporting sheet '{sheet_name}' to PDF: {e}")
            
            workbook.Close(SaveChanges=False)
            workbook = None  # Mark closed to prevent double-close in finally
            if reference_workbook:
                reference_workbook.Close(SaveChanges=False)
                reference_workbook = None
            
            self.last_exported_page_labels = exported_page_labels
            
            # Merge PDFs if multiple
            if not temp_pdf_files:
                utils.logger.error("No PDF files were created")
                return False
            
            if len(temp_pdf_files) > 1:
                success = self.merge_pdfs(temp_pdf_files, output_pdf_path)
            else:
                # Single file - just rename
                shutil.move(temp_pdf_files[0], output_pdf_path)
                temp_pdf_files = []  # Already moved
                success = True
            
            # Cleanup temp files
            for temp_pdf in temp_pdf_files:
                try:
                    if os.path.exists(temp_pdf):
                        os.remove(temp_pdf)
                except:
                    pass
            
            if os.path.exists(output_pdf_path):
                utils.logger.info(f"PDF export successful: {output_pdf_path}")
                return True
            
            return success
            
        except Exception as e:
            utils.logger.error(f"Error in export_sheets_to_pdf: {e}")
            return False
        finally:
            # Close workbook if still open (safety net for exceptions mid-loop)
            if workbook:
                try:
                    workbook.Close(SaveChanges=False)
                except:
                    pass
            if reference_workbook:
                try:
                    reference_workbook.Close(SaveChanges=False)
                except Exception:
                    pass
            # Only cleanup Excel if not managed by the outer retry wrapper
            if not _keep_alive:
                self._cleanup_excel()

    def export_green_sheets_direct(self, file_path, output_pdf_path, print_area="EX1:GR76", _keep_alive=True,
                                   sheet_mode="green"):
        """
        Single-pass high-speed: detects matching sheets and exports them to PDF directly.
        Eliminates opening the workbook twice.
        Returns: (success: bool, green_sheet_names: list)
        """
        colored_sheets = []
        workbook = None
        temp_pdf_files = []
        output_dir = os.path.dirname(output_pdf_path)
        os.makedirs(output_dir, exist_ok=True)
        
        try:
            self._init_excel()
            utils.unblock_file(file_path)
            
            workbook = self.excel_app.Workbooks.Open(
                file_path,
                ReadOnly=True,
                UpdateLinks=False,
                Notify=False,
                AddToMru=False
            )
            
            selected_sheets = []
            for sheet in workbook.Sheets:
                try:
                    if sheet_mode == "visible":
                        is_visible = sheet.Visible in (-1, True, 1)
                        if not is_visible:
                            continue
                        used = sheet.UsedRange
                        if used.Rows.Count == 1 and used.Columns.Count == 1:
                            if used.Value is None or str(used.Value).strip() == "":
                                continue
                        selected_sheets.append(sheet)
                    elif int(sheet.Tab.Color) == config.COLOR_GREEN_TAB:
                        selected_sheets.append(sheet)
                except Exception:
                    continue

            if not selected_sheets:
                description = "visible" if sheet_mode == "visible" else "green-tab"
                utils.logger.warning(f"No {description} sheets found in {file_path}")
                return False, []

            self.last_selected_sheet_names = [sheet.Name.rstrip() for sheet in selected_sheets]
            for idx, sheet in enumerate(selected_sheets, 1):
                sheet_name = sheet.Name.rstrip()
                try:
                    sheet.Activate()
                    ps = sheet.PageSetup
                    ps.PrintArea = print_area
                    ps.PaperSize = 9
                    ps.Zoom = False
                    ps.FitToPagesWide = 1
                    ps.FitToPagesTall = 1
                    ps.CenterHorizontally = True
                except Exception as e:
                    utils.logger.warning(f"PageSetup failed for '{sheet_name}': {e}")
                    
                safe_name = utils.sanitize_filename_strict(sheet_name)
                unique_id = str(uuid.uuid4())[:8]
                temp_pdf = os.path.join(output_dir, f"temp_green_{idx}_{safe_name}_{unique_id}.pdf")
                
                try:
                    sheet.Select()
                    sheet.ExportAsFixedFormat(
                        Type=0,  # xlTypePDF
                        Filename=temp_pdf,
                        Quality=0,
                        IncludeDocProperties=True,
                        IgnorePrintAreas=False,
                        OpenAfterPublish=False
                    )
                    if os.path.exists(temp_pdf):
                        temp_pdf_files.append(temp_pdf)
                        try:
                            import fitz
                            doc = fitz.open(temp_pdf)
                            p_cnt = len(doc)
                            doc.close()
                        except Exception:
                            p_cnt = 1
                        if p_cnt <= 1:
                            colored_sheets.append(sheet_name)
                        else:
                            for p in range(p_cnt):
                                colored_sheets.append(f"{sheet_name}_P{p+1}")
                except Exception as e:
                    utils.logger.error(f"Error exporting green sheet '{sheet_name}' to PDF: {e}")
                    
            workbook.Close(SaveChanges=False)
            workbook = None
            
            if not temp_pdf_files:
                return False, colored_sheets
                
            if len(temp_pdf_files) > 1:
                success = self.merge_pdfs(temp_pdf_files, output_pdf_path)
            else:
                shutil.move(temp_pdf_files[0], output_pdf_path)
                temp_pdf_files = []
                success = True
                
            for temp_pdf in temp_pdf_files:
                try:
                    if os.path.exists(temp_pdf):
                        os.remove(temp_pdf)
                except:
                    pass
                    
            return success, colored_sheets
        except Exception as e:
            utils.logger.error(f"Single-pass green sheet export failed for {file_path}: {e}")
            return False, colored_sheets
        finally:
            if workbook:
                try:
                    workbook.Close(SaveChanges=False)
                except Exception:
                    pass
            if not _keep_alive:
                self._cleanup_excel()

    def export_visible_sheets_direct(self, file_path, output_pdf_path, print_area="J2:BD76", _keep_alive=True):
        """Export visible non-empty sheets while discovering them in the same Excel open."""
        return self.export_green_sheets_direct(
            file_path,
            output_pdf_path,
            print_area=print_area,
            _keep_alive=_keep_alive,
            sheet_mode="visible",
        )
    
    # =========================================================================
    # PDF EXPORT WITH RETRY (LEGACY COMPATIBLE)
    # =========================================================================
    
    def export_sheets_to_pdf_with_retry(self, file_path, sheet_names, output_pdf_path,
                                         print_area="EX1:GR76", max_retries=3, _keep_alive=False,
                                         layout_reference_path=None):
        """
        Export sheets to PDF with retry mechanism and fallback.
        
        Matches legacy export_pdf_with_retry behavior:
        1. Try main method
        2. If failed, try fallback (copy sheets to new workbook)
        3. Retry up to max_retries times
        
        Args:
            file_path: Path to Excel file
            sheet_names: List of sheet names
            output_pdf_path: Output PDF path
            print_area: Print area range
            max_retries: Number of retry attempts
            _keep_alive: If True, keep Excel alive for subsequent calls
            
        Returns:
            True if successful, False otherwise
        """
        import time
        
        # Get safe write path for network files
        safe_output_path, is_fallback = utils.get_safe_write_path(output_pdf_path)
        if is_fallback:
            utils.logger.info(f"Using fallback path: {safe_output_path}")
            output_pdf_path = safe_output_path
        
        try:
            for attempt in range(max_retries):
                try:
                    utils.logger.info(f"PDF export attempt {attempt + 1}/{max_retries}")
                    
                    # === FIX: Keep Excel alive between retries (saves 2-5s startup cost per retry) ===
                    success = self.export_sheets_to_pdf(
                        file_path, sheet_names, output_pdf_path, print_area,
                        _keep_alive=True, layout_reference_path=layout_reference_path
                    )
                    if success and os.path.exists(output_pdf_path):
                        utils.logger.info(f"PDF export successful on attempt {attempt + 1}")
                        return True
                    
                    # If main method failed, try fallback (will reuse self.excel_app)
                    utils.logger.warning(f"Main method failed, trying fallback...")
                    success = self._export_pdf_fallback(file_path, sheet_names, output_pdf_path, print_area)
                    if success and os.path.exists(output_pdf_path):
                        utils.logger.info(f"Fallback export successful")
                        return True
                    
                    utils.logger.warning(f"Attempt {attempt + 1} failed")
                    if attempt < max_retries - 1:
                        time.sleep(2)
                        
                except Exception as e:
                    utils.logger.error(f"Error on attempt {attempt + 1}: {e}")
                    if attempt < max_retries - 1:
                        time.sleep(2)
            
            utils.logger.error(f"All {max_retries} attempts failed")
            return False
        finally:
            if not _keep_alive:
                self._cleanup_excel()
    
    def _export_pdf_fallback(self, file_path, sheet_names, output_pdf_path, print_area="EX1:GR76"):
        """
        Fallback method: Copy sheets to new workbook and export from there.
        
        This handles cases where ExportAsFixedFormat fails on the original file
        (e.g., protected views, network file issues).
        """
        import pythoncom
        import win32com.client
        
        own_excel = False
        workbook = None
        new_workbook = None
        temp_excel_path = None
        excel = None
        
        try:
            temp_excel_path = os.path.join(os.path.dirname(output_pdf_path), f"temp_export_{int(time.time())}.xlsx")
            
            # === FIX: Reuse existing Excel instance to avoid 2-5s startup overhead ===
            if self.excel_app is not None:
                excel = self.excel_app
                utils.logger.debug("Fallback: Reusing existing Excel instance")
            else:
                pythoncom.CoInitialize()
                try:
                    excel = win32com.client.Dispatch("Excel.Application")
                except Exception:
                    excel = win32com.client.DispatchEx("Excel.Application")
                try:
                    excel.Visible = False
                except Exception:
                    pass
                try:
                    excel.DisplayAlerts = False
                except Exception:
                    pass
                try:
                    excel.ScreenUpdating = False
                except Exception:
                    pass
                try:
                    excel.EnableEvents = False
                except Exception:
                    pass
                try:
                    excel.Interactive = False
                    excel.WindowState = -4140  # xlMinimized
                except Exception:
                    pass
                own_excel = True
            
            # Unblock and open source file
            utils.unblock_file(file_path)
            workbook = excel.Workbooks.Open(file_path, ReadOnly=True, UpdateLinks=False)
            
            # Create new workbook
            new_workbook = excel.Workbooks.Add()
            default_sheet_count = new_workbook.Sheets.Count
            
            # Copy sheets in exact order
            copied_count = 0
            for i, sheet_name in enumerate(sheet_names):
                try:
                    source_sheet = None
                    
                    # Try exact match first
                    for sheet in workbook.Sheets:
                        if sheet.Name.rstrip() == sheet_name:
                            source_sheet = sheet
                            break
                    
                    # Fallback: Try variant names (underscore <-> dash)
                    if source_sheet is None:
                        variant1 = sheet_name.replace("-", "_")
                        variant2 = sheet_name.replace("_", "-")
                        
                        for sheet in workbook.Sheets:
                            s_name = sheet.Name.rstrip()
                            if s_name == variant1 or s_name == variant2:
                                source_sheet = sheet
                                utils.logger.info(f"Fallback: Found sheet with variant name: '{s_name}' for '{sheet_name}'")
                                break
                    
                    if source_sheet:
                        source_sheet.Copy(After=new_workbook.Sheets(new_workbook.Sheets.Count))
                        copied_sheet = new_workbook.Sheets(new_workbook.Sheets.Count)
                        copied_count += 1
                        
                        # Setup PageSetup
                        try:
                            ps = copied_sheet.PageSetup
                            ps.PrintArea = print_area
                            ps.Orientation = 1  # Portrait
                            ps.PaperSize = 9    # A4
                            ps.Zoom = False
                            ps.FitToPagesWide = 1
                            ps.FitToPagesTall = 1
                            ps.LeftMargin = 0
                            ps.RightMargin = 0
                            ps.TopMargin = 0
                            ps.BottomMargin = 0
                            ps.HeaderMargin = 0
                            ps.FooterMargin = 0
                            ps.CenterHorizontally = True
                            ps.CenterVertically = True
                            ps.PrintHeadings = False
                            ps.PrintGridlines = False
                        except Exception as ps_err:
                            utils.logger.warning(f"PageSetup error for {sheet_name}: {ps_err}")
                            
                except Exception as e:
                    utils.logger.error(f"Error copying sheet {sheet_name}: {e}")
            
            if copied_count == 0:
                utils.logger.error("Fallback: No sheets were copied")
                return False
                
            # Remove default empty sheets
            for _ in range(default_sheet_count):
                try:
                    if new_workbook.Sheets.Count > copied_count:
                        new_workbook.Sheets(1).Delete()
                except Exception:
                    break
            
            # Save temp file
            new_workbook.SaveAs(temp_excel_path)
            
            # Export to PDF
            new_workbook.ExportAsFixedFormat(
                Type=0,  # xlTypePDF
                Filename=output_pdf_path,
                Quality=0,
                IncludeDocProperties=True,
                IgnorePrintAreas=False,
                OpenAfterPublish=False
            )
            
            utils.logger.info(f"Fallback export completed: {output_pdf_path}")
            return os.path.exists(output_pdf_path)
            
        except Exception as e:
            utils.logger.error(f"Fallback export failed: {e}")
            return False
        finally:
            try:
                if new_workbook:
                    new_workbook.Close(SaveChanges=False)
            except:
                pass
            try:
                if workbook:
                    workbook.Close(SaveChanges=False)
            except:
                pass
            # === FIX: Only quit Excel if we created it, don't kill caller's instance ===
            if own_excel:
                try:
                    if excel:
                        excel.Quit()
                except:
                    pass
                pythoncom.CoUninitialize()
            try:
                if temp_excel_path and os.path.exists(temp_excel_path):
                    os.remove(temp_excel_path)
            except:
                pass
    
    # =========================================================================
    # PDF MERGING
    # =========================================================================
    
    def merge_pdfs(self, pdf_list, output_path):
        """
        Merge multiple PDF files into one.
        
        Args:
            pdf_list: List of PDF file paths to merge
            output_path: Output merged PDF path
            
        Returns:
            True if successful, False otherwise
        """
        if not PYPDF2_AVAILABLE:
            utils.logger.error("PyPDF2 not available for merging PDFs")
            # Fallback: copy first file
            if pdf_list:
                shutil.copy2(pdf_list[0], output_path)
                return True
            return False
        
        try:
            merger = PdfMerger()
            
            for pdf_path in pdf_list:
                if os.path.exists(pdf_path):
                    merger.append(pdf_path)
            
            merger.write(output_path)
            merger.close()
            
            utils.logger.info(f"Merged {len(pdf_list)} PDFs into {output_path}")
            return True
            
        except Exception as e:
            utils.logger.error(f"Error merging PDFs: {e}")
            # Fallback: copy first file
            if pdf_list and os.path.exists(pdf_list[0]):
                shutil.copy2(pdf_list[0], output_path)
                return True
            return False
    
    # =========================================================================
    # PDF TO IMAGE RENDERING (PyMuPDF)
    # =========================================================================
    
    def render_pdf_to_images(self, pdf_path, dpi=100):
        """
        Render PDF pages to PIL Image objects using PyMuPDF.
        
        Args:
            pdf_path: Path to PDF file
            dpi: Resolution for rendering (default: 100)
            
        Returns:
            List of PIL Image objects, one per page
        """
        images = []
        
        if not PYMUPDF_AVAILABLE:
            utils.logger.error("PyMuPDF not available for PDF rendering")
            return images
        
        try:
            # Calculate zoom based on DPI (72 is PDF default)
            zoom = dpi / 72.0
            matrix = fitz.Matrix(zoom, zoom)
            
            doc = fitz.open(pdf_path)
            
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                pix = page.get_pixmap(matrix=matrix, alpha=False)
                
                # Convert to PIL Image
                img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
                images.append(img)
            
            doc.close()
            utils.logger.info(f"Rendered {len(images)} pages from {pdf_path} at DPI={dpi}")
            
        except Exception as e:
            utils.logger.error(f"Error rendering PDF to images: {e}")
        
        return images
    
    def render_pdf_page(self, pdf_path, page_num=0, dpi=100):
        """
        Render a single PDF page to PIL Image.
        
        Args:
            pdf_path: Path to PDF file
            page_num: Page number (0-indexed)
            dpi: Resolution for rendering
            
        Returns:
            PIL Image or None if failed
        """
        if not PYMUPDF_AVAILABLE:
            utils.logger.error("PyMuPDF not available")
            return None
        
        try:
            zoom = dpi / 72.0
            matrix = fitz.Matrix(zoom, zoom)
            
            doc = fitz.open(pdf_path)
            if page_num >= len(doc):
                utils.logger.error(f"Page {page_num} does not exist in {pdf_path}")
                doc.close()
                return None
            
            page = doc.load_page(page_num)
            pix = page.get_pixmap(matrix=matrix, alpha=False)
            img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
            
            doc.close()
            return img
            
        except Exception as e:
            utils.logger.error(f"Error rendering PDF page: {e}")
            return None
    
    # =========================================================================
    # PDF COMPARISON WITH HIGHLIGHTING
    # =========================================================================
    
    def compare_pdfs_and_highlight(self, pdf_new_path, pdf_old_path, output_path, 
                                    sheet_name="comparison",
                                    dpi=100,
                                    diff_threshold=40,
                                    dilate_size=3,
                                    dilate_iterations=2,
                                    highlight_color="#ff0000",
                                    fill_opacity=40):
        """
        Compare two PDF files and create a side-by-side comparison PDF with highlights.
        
        === OPTIMIZED VERSION (Cấp độ 3) ===
        - Sử dụng OpenCV thay PIL cho xử lý ảnh (nhanh hơn 10-50x)
        - Short-circuit: Bỏ qua xử lý nặng nếu ảnh giống nhau 100%
        
        Args:
            pdf_new_path: Path to new PDF
            pdf_old_path: Path to old PDF
            output_path: Output comparison PDF path
            sheet_name: Name for labeling
            dpi: Rendering DPI
            diff_threshold: Pixel difference threshold (0-255)
            dilate_size: Dilation kernel size (odd number 1-9)
            dilate_iterations: Number of dilation passes (1-5)
            highlight_color: Hex color for highlight overlay
            fill_opacity: Opacity percentage (0-100)
            
        Returns:
            Output path if successful, None if failed
        """
        if not PYMUPDF_AVAILABLE:
            utils.logger.error("PyMuPDF required for PDF comparison")
            return None
        
        try:
            # Render both PDFs to images
            img_new = self.render_pdf_page(pdf_new_path, page_num=0, dpi=dpi)
            img_old = self.render_pdf_page(pdf_old_path, page_num=0, dpi=dpi)
            
            if img_new is None or img_old is None:
                utils.logger.error("Failed to render PDF pages")
                return None
            
            # === OPTIMIZATION: Short-circuit for identical images ===
            if OPENCV_COMPARE_AVAILABLE:
                if quick_compare_hash(img_new, img_old):
                    utils.logger.info(f"[OPTIMIZED] Images identical, skipping heavy processing for {sheet_name}")
                    # Still create side-by-side but without expensive processing
                    left_img = img_old.convert('RGB')
                    right_img = img_new.convert('RGB')
                    side_by_side = create_side_by_side(left_img, right_img)
                    
                    if REPORTLAB_AVAILABLE:
                        self._create_pdf_with_reportlab(side_by_side, output_path)
                    else:
                        side_by_side.save(output_path, "PDF", resolution=float(dpi))
                    
                    return output_path
            
            # === MAIN COMPARISON LOGIC ===
            if OPENCV_COMPARE_AVAILABLE:
                # Use OpenCV-optimized comparison (10-50x faster)
                utils.logger.debug(f"[OPTIMIZED] Using OpenCV for image comparison: {sheet_name}")
                left_img, right_img = compare_images_opencv(
                    img_new, img_old,
                    diff_threshold=diff_threshold,
                    dilate_size=dilate_size,
                    dilate_iterations=dilate_iterations,
                    highlight_color=highlight_color,
                    fill_opacity=fill_opacity
                )
                side_by_side = create_side_by_side(left_img, right_img)
            else:
                # Fallback: Original PIL-based comparison
                utils.logger.debug(f"Using PIL for image comparison (OpenCV not available): {sheet_name}")
                
                # Resize if different sizes
                if img_new.size != img_old.size:
                    img_old = img_old.resize(img_new.size, Image.Resampling.LANCZOS)
                
                # Compute difference
                diff = ImageChops.difference(img_new.convert('RGB'), img_old.convert('RGB'))
                diff_gray = diff.convert('L')
                
                # Apply threshold
                threshold = max(0, min(255, diff_threshold))
                mask = diff_gray.point(lambda x: 255 if x > threshold else 0)
                
                # Dilate mask
                if dilate_size % 2 == 0:
                    dilate_size = max(1, dilate_size - 1)
                dilate_size = max(1, min(9, dilate_size))
                dilate_iterations = max(1, min(3, dilate_iterations))
                
                for _ in range(dilate_iterations):
                    mask = mask.filter(ImageFilter.MaxFilter(dilate_size))
                
                # Create highlight overlay
                opacity = max(0, min(100, fill_opacity))
                rgba_color = self._hex_to_rgba(highlight_color, opacity)
                
                red_overlay = Image.new('RGBA', img_new.size, rgba_color)
                mask_l = mask.convert('L')
                overlay_masked = Image.new('RGBA', img_new.size, (0, 0, 0, 0))
                overlay_masked.paste(red_overlay, (0, 0), mask_l)
                
                # Apply overlay to new image
                combined_new = Image.alpha_composite(img_new.convert('RGBA'), overlay_masked)
                
                # Create side-by-side
                left_img = img_old.convert('RGB')
                right_img = combined_new.convert('RGB')
                
                side_width = left_img.width + right_img.width
                side_height = max(left_img.height, right_img.height)
                
                side_by_side = Image.new('RGB', (side_width, side_height), (255, 255, 255))
                side_by_side.paste(left_img, (0, 0))
                side_by_side.paste(right_img, (left_img.width, 0))
            
            # Create PDF output
            if REPORTLAB_AVAILABLE:
                self._create_pdf_with_reportlab(side_by_side, output_path)
            else:
                side_by_side.save(output_path, "PDF", resolution=float(dpi))
            
            utils.logger.info(f"Comparison PDF created: {output_path}")
            return output_path
            
        except Exception as e:
            utils.logger.error(f"Error comparing PDFs: {e}")
            return None

    
    def _create_pdf_with_reportlab(self, image, output_path):
        """Create landscape PDF from image using reportlab."""
        page_size = landscape(A4)
        page_width, page_height = page_size
        
        c = canvas.Canvas(output_path, pagesize=page_size)
        
        # Fit image to page
        img_width, img_height = image.size
        scale = min(page_width / img_width, page_height / img_height) * 0.98
        
        new_w = img_width * scale
        new_h = img_height * scale
        x_offset = (page_width - new_w) / 2
        y_offset = (page_height - new_h) / 2
        
        c.drawImage(ImageReader(image), x_offset, y_offset, new_w, new_h)
        c.save()
    
    def _hex_to_rgba(self, hex_color, opacity=100):
        """Convert hex color + opacity to RGBA tuple."""
        hex_color = hex_color.lstrip('#')
        r, g, b = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        alpha = int(255 * opacity / 100)
        return (r, g, b, alpha)
    
    # =========================================================================
    # LEGACY COMPATIBILITY: convert_pdf_to_images (saves to files)
    # =========================================================================
    
    def convert_pdf_to_images(self, pdf_path, output_folder, dpi=100):
        """
        Converts PDF pages to image files (legacy compatibility).
        
        Args:
            pdf_path: Path to PDF file
            output_folder: Folder to save images
            dpi: Resolution
            
        Returns:
            List of saved image paths
        """
        image_paths = []
        
        # Try PyMuPDF first
        if PYMUPDF_AVAILABLE:
            images = self.render_pdf_to_images(pdf_path, dpi)
            for i, img in enumerate(images):
                image_path = os.path.join(output_folder, f"page_{i}.png")
                img.save(image_path, "PNG")
                image_paths.append(image_path)
            return image_paths
        
        # Fallback to pdf2image
        if PDF2IMAGE_AVAILABLE and convert_from_path:
            try:
                images = convert_from_path(pdf_path, dpi=dpi)
                for i, image in enumerate(images):
                    image_path = os.path.join(output_folder, f"page_{i}.png")
                    image.save(image_path, "PNG")
                    image_paths.append(image_path)
                return image_paths
            except Exception as e:
                utils.logger.error(f"PDF conversion with pdf2image failed: {e}")
        
        utils.logger.error("No PDF rendering library available")
        return image_paths
    
    @staticmethod
    def export_sheet_to_pdf(sheet, output_path, excel_app):
        """
        Exports a specific sheet to PDF (legacy compatibility).
        """
        try:
            sheet.ExportAsFixedFormat(0, output_path)  # 0 = xlTypePDF
            return True
        except Exception as e:
            utils.logger.error(f"Export to PDF failed: {e}")
            return False

    def merge_all_result_pdfs(self, base_path, output_filename="comparison_ALL_SHEETS.pdf", dpi=100):
        """
        Tìm tất cả file comparison_*.pdf và ảnh comparison_*.png/jpg,
        gộp chúng thành một file PDF duy nhất.
        
        Args:
            base_path: Thư mục chứa các file cần gộp
            output_filename: Tên file output
            dpi: DPI khi chuyển ảnh sang PDF
            
        Returns:
            Đường dẫn file gộp nếu thành công, None nếu thất bại
        """
        import glob
        
        if not os.path.exists(base_path):
            return None
        
        # Tìm tất cả PDF comparison
        pdf_patterns = ["comparison_*.pdf"]
        files = []
        for pattern in pdf_patterns:
            files.extend(sorted(glob.glob(os.path.join(base_path, pattern))))
        
        # Loại trừ file output nếu đã tồn tại
        merged_path = os.path.join(base_path, output_filename)
        files = [f for f in files if os.path.abspath(f) != os.path.abspath(merged_path)]
        
        # Chuyển ảnh comparison thành PDF tạm
        img_patterns = ["comparison_*.png", "comparison_*.jpg", "*_so_sanh.png", "*_so_sanh.jpg"]
        temp_pdfs = []
        
        for pattern in img_patterns:
            for img_path in sorted(glob.glob(os.path.join(base_path, pattern))):
                try:
                    img = Image.open(img_path).convert('RGB')
                    temp_pdf = os.path.splitext(img_path)[0] + "_imgconv.pdf"
                    img.save(temp_pdf, "PDF", resolution=float(dpi))
                    temp_pdfs.append(temp_pdf)
                    files.append(temp_pdf)
                except Exception as e:
                    utils.logger.debug(f"Cannot convert image to PDF: {img_path} - {e}")
        
        if not files:
            utils.logger.info("No comparison files found to merge")
            return None
        
        # Gộp PDF
        if not PYPDF2_AVAILABLE:
            utils.logger.error("PyPDF2 not available for merging")
            return None
        
        try:
            merger = PyPDF2.PdfMerger()
            added = 0
            
            for f in files:
                try:
                    if not os.path.exists(f):
                        continue
                    
                    # Kiểm tra PDF hợp lệ
                    with open(f, 'rb') as fh:
                        try:
                            reader = PyPDF2.PdfReader(fh)
                            _ = len(reader.pages)
                        except Exception:
                            utils.logger.debug(f"Invalid PDF skipped: {f}")
                            continue
                    
                    merger.append(f)
                    added += 1
                    
                except Exception as e:
                    utils.logger.debug(f"Error adding {f}: {e}")
            
            if added == 0:
                merger.close()
                return None
            
            with open(merged_path, 'wb') as fout:
                merger.write(fout)
            merger.close()
            
            utils.logger.info(f"Merged {added} PDFs to: {output_filename}")
            
            # Cleanup temp PDFs
            for temp_pdf in temp_pdfs:
                try:
                    if os.path.exists(temp_pdf):
                        os.remove(temp_pdf)
                except:
                    pass
            
            return merged_path
            
        except Exception as e:
            utils.logger.error(f"Merge failed: {e}")
            return None

