"""
Report Service - Tạo báo cáo Excel kết quả so sánh

Enhanced để hỗ trợ:
- Embedded images từ PDF preview
- Hyperlinks đến file PDF/ảnh
- Multi-sheet report cho từng kết quả
"""

import os
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as xlImage
import utils

# Check PyMuPDF availability
PYMUPDF_AVAILABLE = False
try:
    import fitz
    PYMUPDF_AVAILABLE = True
except ImportError:
    pass


class ReportService:
    """Service tạo báo cáo Excel với kết quả so sánh."""
    
    def __init__(self, output_folder):
        self.output_folder = output_folder
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Kết quả so sánh"
        self._init_header()
        self.current_row = 2
        self.temp_preview_files = []  # Track temp files for cleanup
    
    def _init_header(self):
        """Khởi tạo header cho sheet tổng hợp."""
        headers = ["STT", "Tên File Mới", "Tên Sheet", "Trạng thái", "Chi tiết", "Link Ảnh Sai Khác"]
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        # Column widths
        self.ws.column_dimensions['A'].width = 5
        self.ws.column_dimensions['B'].width = 30
        self.ws.column_dimensions['C'].width = 20
        self.ws.column_dimensions['D'].width = 15
        self.ws.column_dimensions['E'].width = 40
        self.ws.column_dimensions['F'].width = 50

    def add_result(self, file_name, sheet_name, status, detail, diff_image_path=None):
        """
        Thêm một dòng kết quả vào báo cáo.
        
        Args:
            file_name: Tên file được so sánh
            sheet_name: Tên sheet
            status: Trạng thái (OK, FAIL, ERROR, WARNING, MISSING)
            detail: Chi tiết kết quả
            diff_image_path: Đường dẫn ảnh sai khác (nếu có)
        """
        row = self.current_row
        
        # Status colors
        status_colors = {
            "OK": "00B050",      # Green
            "FAIL": "FF0000",    # Red
            "ERROR": "FF6600",   # Orange
            "WARNING": "FFD700", # Yellow
            "MISSING": "808080"  # Gray
        }
        status_color = status_colors.get(status, "000000")
        
        self.ws.cell(row=row, column=1, value=row-1)  # STT
        self.ws.cell(row=row, column=2, value=file_name)
        self.ws.cell(row=row, column=3, value=sheet_name)
        
        cell_status = self.ws.cell(row=row, column=4, value=status)
        cell_status.font = Font(color=status_color, bold=True)
        
        self.ws.cell(row=row, column=5, value=detail)
        
        if diff_image_path and os.path.exists(diff_image_path):
            cell_link = self.ws.cell(row=row, column=6, value="Mở ảnh sai khác")
            cell_link.hyperlink = diff_image_path
            cell_link.font = Font(color="0000FF", underline="single")
        elif status not in ("OK", "MISSING"):
            self.ws.cell(row=row, column=6, value="Không có ảnh")

        self.current_row += 1

    def add_comparison_sheet(self, comparison_result_path, sheet_name=None, embed_preview=True, dpi=100):
        """
        Thêm một sheet mới với kết quả so sánh (embedded image hoặc hyperlink).
        
        Args:
            comparison_result_path: Đường dẫn file PDF hoặc ảnh kết quả
            sheet_name: Tên sheet (tự động nếu None)
            embed_preview: Nếu True, chèn ảnh preview; nếu False, chỉ hyperlink
            dpi: DPI cho render preview
            
        Returns:
            True nếu thành công
        """
        if not os.path.exists(comparison_result_path):
            return False
        
        # Generate sheet name
        if not sheet_name:
            filename = os.path.basename(comparison_result_path)
            if filename.startswith("comparison_"):
                sheet_name = filename.replace("comparison_", "").rsplit(".", 1)[0]
            else:
                sheet_name = filename.rsplit(".", 1)[0]
        
        # Sanitize sheet name (max 31 chars)
        safe_sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in '_- ')[:31]
        if not safe_sheet_name:
            safe_sheet_name = f"Result_{self.wb.worksheets.__len__()}"
        
        try:
            ws = self.wb.create_sheet(title=safe_sheet_name)
        except Exception:
            ws = self.wb.create_sheet(title=f"Sheet_{self.wb.worksheets.__len__()}")
        
        # Handle different file types
        if comparison_result_path.endswith('.txt'):
            ws['A1'] = f"Không có khác biệt trong: {sheet_name}"
            ws['A2'] = "Nội dung giống hệt nhau."
            return True
        
        if comparison_result_path.endswith('.pdf') and embed_preview and PYMUPDF_AVAILABLE:
            return self._embed_pdf_preview(ws, comparison_result_path, dpi)
        
        if comparison_result_path.endswith(('.png', '.jpg', '.jpeg')):
            return self._embed_image(ws, comparison_result_path)
        
        # Fallback: hyperlink
        return self._add_hyperlink(ws, comparison_result_path)
    
    def _embed_pdf_preview(self, ws, pdf_path, dpi=100):
        """Chèn preview của trang đầu PDF vào worksheet."""
        try:
            from PIL import Image as PILImage
            
            zoom = dpi / 72.0
            mtx = fitz.Matrix(zoom, zoom)
            
            with fitz.open(pdf_path) as doc:
                page = doc.load_page(0)
                pm = page.get_pixmap(matrix=mtx, alpha=False)
                pil_img = PILImage.frombytes("RGB", (pm.width, pm.height), pm.samples)
                
                # Save temp preview
                safe_name = "".join(c for c in os.path.basename(pdf_path) if c.isalnum() or c in '_-')
                temp_path = os.path.join(self.output_folder, f"temp_preview_{safe_name}_{int(time.time()*1000)}.png")
                pil_img.save(temp_path, 'PNG')
                
                if os.path.exists(temp_path):
                    img = xlImage(temp_path)
                    ws.add_image(img, 'A1')
                    self.temp_preview_files.append(temp_path)
                    return True
                    
        except Exception as e:
            utils.logger.debug(f"PDF preview failed: {e}")
        
        # Fallback to hyperlink
        return self._add_hyperlink(ws, pdf_path)
    
    def _embed_image(self, ws, image_path):
        """Chèn ảnh vào worksheet."""
        try:
            img = xlImage(image_path)
            ws.add_image(img, 'A1')
            return True
        except Exception as e:
            utils.logger.debug(f"Image embed failed: {e}")
            return self._add_hyperlink(ws, image_path)
    
    def _add_hyperlink(self, ws, file_path):
        """Thêm hyperlink đến file."""
        try:
            from pathlib import Path
            
            cell = ws['A1']
            ext = os.path.splitext(file_path)[1].lower()
            
            if ext == '.pdf':
                cell.value = "Mở PDF so sánh"
            elif ext in ('.png', '.jpg', '.jpeg'):
                cell.value = "Mở ảnh so sánh"
            else:
                cell.value = "Mở file kết quả"
            
            cell.hyperlink = Path(file_path).resolve().as_uri()
            cell.font = Font(color="0000FF", underline="single")
            return True
            
        except Exception as e:
            utils.logger.debug(f"Hyperlink failed: {e}")
            ws['A1'] = f"Xem: {os.path.basename(file_path)}"
            return False

    def create_from_pdf_results(self, comparison_results, embed_previews=True, dpi=100):
        """
        Tạo báo cáo nhiều sheet từ danh sách kết quả so sánh PDF.
        
        Args:
            comparison_results: List đường dẫn file PDF/ảnh kết quả
            embed_previews: Chèn preview hay chỉ hyperlink
            dpi: DPI cho render preview
        """
        for idx, result_path in enumerate(comparison_results, 1):
            self.add_comparison_sheet(result_path, embed_preview=embed_previews, dpi=dpi)

    @staticmethod
    def create_per_file_excel_result(new_file_path, comparison_images, old_images, output_folder, 
                                      is_pdf_method=False, dpi=100):
        """
        Tạo file Excel kết quả riêng cho mỗi cặp file theo format legacy.
        
        Legacy format:
        - Tên file: "Kết quả_{tên_file_mới}.xlsx" hoặc "Kết quả_PDF_{tên_file_mới}.xlsx"
        - Mỗi sheet chứa ảnh so sánh (cũ bên trái, mới bên phải hoặc side-by-side)
        
        Args:
            new_file_path: Đường dẫn file Excel mới (dùng để lấy tên)
            comparison_images: Dict {sheet_name: comparison_image_path} - ảnh so sánh side-by-side
            old_images: Dict {sheet_name: old_image_path} - ảnh file cũ (nếu có)
            output_folder: Thư mục lưu kết quả
            is_pdf_method: True nếu dùng PDF method
            dpi: DPI cho render (dùng với PDF)
            
        Returns:
            Đường dẫn file Excel kết quả nếu thành công, None nếu thất bại
        """
        try:
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as xlImage
            
            excel_filename = os.path.basename(new_file_path)
            excel_name_only = os.path.splitext(excel_filename)[0]
            
            if is_pdf_method:
                result_filename = f"Kết quả_PDF_{excel_name_only}.xlsx"
            else:
                result_filename = f"Kết quả_{excel_name_only}.xlsx"
            
            result_path = os.path.join(output_folder, result_filename)
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            
            # Nếu không có comparison images, tạo sheet placeholder
            if not comparison_images:
                ws = wb.create_sheet(title="Không có kết quả")
                ws['A1'] = "Không tìm thấy ảnh kết quả so sánh"
                ws['A2'] = "Có thể do không có sheet nào có màu tab đúng hoặc không có sự khác biệt."
                wb.save(result_path)
                utils.logger.info(f"Empty result saved: {result_path}")
                return result_path
            
            # Tạo sheet cho mỗi ảnh so sánh
            for sheet_name, comparison_path in comparison_images.items():
                # Sanitize sheet name
                safe_sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in '_- ')[:30]
                if not safe_sheet_name:
                    safe_sheet_name = f"Result_{wb.worksheets.__len__()+1}"
                
                try:
                    ws = wb.create_sheet(title=f" {safe_sheet_name}")
                except Exception:
                    ws = wb.create_sheet(title=f"Sheet_{wb.worksheets.__len__()+1}")
                
                current_col = 'A'
                
                # Chèn ảnh cũ nếu có
                old_img_path = old_images.get(sheet_name) if old_images else None
                if old_img_path and os.path.exists(old_img_path):
                    try:
                        old_img = xlImage(old_img_path)
                        ws.add_image(old_img, 'A1')
                        
                        # Tính vị trí cột tiếp theo dựa trên độ rộng ảnh
                        from PIL import Image as PILImage
                        with PILImage.open(old_img_path) as pil_img:
                            # Excel column width: ~7.5 pixels per unit
                            img_cols = int(pil_img.width / 7.5) + 1
                            current_col = chr(ord('A') + img_cols)
                            if img_cols > 25:
                                current_col = 'Z'
                    except Exception as e:
                        utils.logger.warning(f"Could not embed old image: {e}")
                        current_col = 'G'  # Default position
                
                # Chèn ảnh so sánh
                if comparison_path and os.path.exists(comparison_path):
                    try:
                        if comparison_path.endswith('.pdf') and PYMUPDF_AVAILABLE:
                            # Render PDF to image first
                            from PIL import Image as PILImage
                            zoom = dpi / 72.0
                            mtx = fitz.Matrix(zoom, zoom)
                            
                            with fitz.open(comparison_path) as doc:
                                page = doc.load_page(0)
                                pm = page.get_pixmap(matrix=mtx, alpha=False)
                                pil_img = PILImage.frombytes("RGB", (pm.width, pm.height), pm.samples)
                                
                                temp_preview = os.path.join(output_folder, f"__temp_preview_{sheet_name}_{int(time.time())}.png")
                                pil_img.save(temp_preview, 'PNG')
                                
                                if os.path.exists(temp_preview):
                                    comp_img = xlImage(temp_preview)
                                    ws.add_image(comp_img, f'{current_col}1')
                                    # Cleanup temp file after save (done in finally)
                        else:
                            # Direct image
                            comp_img = xlImage(comparison_path)
                            ws.add_image(comp_img, f'{current_col}1')
                            
                    except Exception as e:
                        utils.logger.warning(f"Could not embed comparison image: {e}")
                        ws['A1'] = f"Xem ảnh: {comparison_path}"
            
            # Save workbook
            wb.save(result_path)
            utils.logger.info(f"Per-file result saved: {result_path}")
            return result_path
            
        except Exception as e:
            utils.logger.error(f"Error creating per-file Excel result: {e}")
            return None

    def save(self, filename="TongHopKetQua.xlsx"):
        """
        Lưu file báo cáo.
        
        Args:
            filename: Tên file báo cáo
            
        Returns:
            Đường dẫn file nếu thành công, None nếu thất bại
        """
        report_path = os.path.join(self.output_folder, filename)
        
        try:
            self.wb.save(report_path)
            utils.logger.info(f"Report saved: {report_path}")
            
            # Cleanup temp preview files
            for temp_file in self.temp_preview_files:
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                except:
                    pass
            self.temp_preview_files.clear()
            
            return report_path
            
        except Exception as e:
            utils.logger.error(f"Failed to save report: {e}")
            return None


class LibraryValidator:
    """
    Kiểm tra và xác nhận các thư viện cần thiết.
    """
    
    REQUIRED_LIBRARIES = {
        'openpyxl': 'Tạo file Excel',
        'PIL': 'Xử lý ảnh',
        'win32com': 'Điều khiển Excel COM',
    }
    
    OPTIONAL_LIBRARIES = {
        'fitz': 'PyMuPDF - Render PDF (khuyến nghị)',
        'PyPDF2': 'Gộp PDF',
        'reportlab': 'Tạo PDF',
        'pyautogui': 'Chụp screenshot',
        'pdf2image': 'Chuyển PDF sang ảnh (backup)',
    }
    
    @staticmethod
    def check_all():
        """
        Kiểm tra tất cả thư viện.
        
        Returns:
            Dict với trạng thái các thư viện
        """
        results = {
            'required': {},
            'optional': {},
            'missing_required': [],
            'missing_optional': [],
            'all_required_available': True
        }
        
        # Check required
        for lib, desc in LibraryValidator.REQUIRED_LIBRARIES.items():
            available = LibraryValidator._check_library(lib)
            results['required'][lib] = {'available': available, 'description': desc}
            if not available:
                results['missing_required'].append(lib)
                results['all_required_available'] = False
        
        # Check optional
        for lib, desc in LibraryValidator.OPTIONAL_LIBRARIES.items():
            available = LibraryValidator._check_library(lib)
            results['optional'][lib] = {'available': available, 'description': desc}
            if not available:
                results['missing_optional'].append(lib)
        
        return results
    
    @staticmethod
    def _check_library(lib_name):
        """Kiểm tra một thư viện có khả dụng không."""
        try:
            if lib_name == 'PIL':
                from PIL import Image
            elif lib_name == 'win32com':
                import win32com.client
            elif lib_name == 'fitz':
                import fitz
            else:
                __import__(lib_name)
            return True
        except ImportError:
            return False
    
    @staticmethod
    def get_status_message():
        """
        Lấy thông báo trạng thái thư viện.
        
        Returns:
            Tuple (message, is_ok)
        """
        results = LibraryValidator.check_all()
        
        if results['all_required_available']:
            if not results['missing_optional']:
                return "✅ Tất cả thư viện đã sẵn sàng", True
            else:
                missing = ", ".join(results['missing_optional'])
                return f"✅ OK (thiếu tùy chọn: {missing})", True
        else:
            missing = ", ".join(results['missing_required'])
            return f"❌ Thiếu thư viện bắt buộc: {missing}", False
    
    @staticmethod
    def print_status():
        """In trạng thái tất cả thư viện ra console."""
        results = LibraryValidator.check_all()
        
        print("\n=== KIỂM TRA THƯ VIỆN ===")
        print("\nBẮT BUỘC:")
        for lib, info in results['required'].items():
            status = "✅" if info['available'] else "❌"
            print(f"  {status} {lib}: {info['description']}")
        
        print("\nTÙY CHỌN:")
        for lib, info in results['optional'].items():
            status = "✅" if info['available'] else "⚠️"
            print(f"  {status} {lib}: {info['description']}")
        
        print("")
        msg, ok = LibraryValidator.get_status_message()
        print(msg)
        return ok
