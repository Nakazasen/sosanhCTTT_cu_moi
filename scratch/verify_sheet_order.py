import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

f1 = r"\\fstvn01\Data\00_KDTVN Common(KDTVN共通)\⑤Production Engineering(製造技術)\Hang muc can luu\Vinh\sosanhDUKC_CTTT\KetQuaSoSanh_CTTT_20260826_194817\Kết quả_PDF_DUKC mới.xlsx"
wb1 = openpyxl.load_workbook(f1)
print("=== File 1: DUKC mới.xlsx ===")
print("Sheet order:", wb1.sheetnames)

f2 = r"\\fstvn01\Data\00_KDTVN Common(KDTVN共通)\⑤Production Engineering(製造技術)\Hang muc can luu\Vinh\sosanhDUKC_CTTT\KetQuaSoSanh_CTTT_20260826_194901\Kết quả_PDF_【緊急対応添付資料】3V2RV00280_EXIT_ASSY_23.01.2025_1 Mới.xlsx"
wb2 = openpyxl.load_workbook(f2)
print("\n=== File 2: 【緊急対応添付資料】...xlsx ===")
print("Sheet order:", wb2.sheetnames)
