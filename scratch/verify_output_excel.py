import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

f2 = r"\\fstvn01\Data\00_KDTVN Common(KDTVN共通)\⑤Production Engineering(製造技術)\Hang muc can luu\Vinh\sosanhDUKC_CTTT\KetQuaSoSanh_CTTT_20260826_194309\Kết quả_PDF_【緊急対応添付資料】3V2RV00280_EXIT_ASSY_23.01.2025_1 Mới.xlsx"
wb2 = openpyxl.load_workbook(f2)
print("=== File: 【緊急対応添付資料】...xlsx ===")
print("Sheets in Excel:", wb2.sheetnames)
for sname in wb2.sheetnames:
    ws = wb2[sname]
    print(f'  Sheet "{sname}": max_row={ws.max_row}, max_column={ws.max_column}, images={len(getattr(ws, "_images", []))}')
