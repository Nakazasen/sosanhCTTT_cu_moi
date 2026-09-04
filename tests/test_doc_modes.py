import unittest
from unittest.mock import MagicMock, patch
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import config
from ui.translations import get_text, LANGUAGES
from services.pdf_service import PDFService
from services.settings_service import SettingsService
from services.validation_service import ValidationService
from core.comparator import Comparator
from ui.main_window_modern import CustomModeConfigDialog, MainWindow


class TestDocModesConfigAndTranslations(unittest.TestCase):
    """Test configuration constants and translations for document modes."""

    def test_doc_mode_constants_exist(self):
        self.assertEqual(config.DOC_MODE_STANDARD_CTTT, "standard_cttt")
        self.assertEqual(config.DOC_MODE_DUKC_CTTT, "dukc_cttt")
        self.assertEqual(config.DOC_MODE_DUKC_OTHER, "dukc_other")
        self.assertEqual(config.DOC_MODE_CUSTOM, "custom")

        self.assertEqual(config.PRINT_AREA_STANDARD_CTTT, "EX1:GR76")
        self.assertEqual(config.PRINT_AREA_DUKC_CTTT, "J2:BD76")
        self.assertEqual(config.PRINT_AREA_DUKC_OTHER, "A1:AT120")
        self.assertEqual(config.DEFAULT_CUSTOM_PRINT_AREA, "A1:Z100")

    def test_default_output_folder_name_exists(self):
        self.assertEqual(config.DEFAULT_OUTPUT_FOLDER_NAME, "KetQuaSoSanh_CTTT")

    def test_translations_for_all_modes_and_languages(self):
        mode_keys = [
            "doc_type_label", "mode_standard_cttt", "mode_dukc_cttt", "mode_dukc_other", "mode_custom",
            "print_area_label", "custom_config_title", "custom_range_group", "custom_sheet_group"
        ]
        for lang in LANGUAGES.keys():
            for key in mode_keys:
                text = get_text(key, lang)
                self.assertIsNotNone(text, f"Missing translation for {key} in {lang}")
                self.assertTrue(len(text) > 0, f"Empty translation for {key} in {lang}")


class TestVisibleSheetsAndMatching(unittest.TestCase):
    """Test visible sheets extraction and sheet matching logic."""
    
    def test_get_visible_sheets_mock(self):
        pdf_service = PDFService()
        pdf_service.excel_app = MagicMock()

        # Mock Workbook with 4 sheets: 2 visible with data, 1 visible empty, 1 hidden
        sheet1 = MagicMock()
        sheet1.Name = "SheetA "
        sheet1.Visible = -1
        sheet1.UsedRange.Rows.Count = 10
        sheet1.UsedRange.Columns.Count = 5
        sheet1.UsedRange.Value = "Data"

        sheet2 = MagicMock()
        sheet2.Name = "SheetB"
        sheet2.Visible = -1
        sheet2.UsedRange.Rows.Count = 1
        sheet2.UsedRange.Columns.Count = 1
        sheet2.UsedRange.Value = ""  # Empty sheet

        sheet3 = MagicMock()
        sheet3.Name = "data1"
        sheet3.Visible = 0  # Hidden sheet

        sheet4 = MagicMock()
        sheet4.Name = "SheetC"
        sheet4.Visible = -1
        sheet4.UsedRange.Rows.Count = 50
        sheet4.UsedRange.Columns.Count = 20
        sheet4.UsedRange.Value = "More Data"
        
        mock_wb = MagicMock()
        mock_wb.Sheets = [sheet1, sheet2, sheet3, sheet4]
        pdf_service.excel_app.Workbooks.Open.return_value = mock_wb
        
        visible = pdf_service.get_visible_sheets("dummy.xlsx")
        
        # Should contain SheetA and SheetC only (skips empty SheetB and hidden data1)
        self.assertEqual(visible, ["SheetA", "SheetC"])

    def test_get_visible_sheets_fast_uses_workbook_metadata(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, "visible.xlsx")
            workbook = Workbook()
            workbook.active.title = "Visible"
            workbook.active["A1"] = "data"
            workbook.create_sheet("Empty visible")
            hidden = workbook.create_sheet("Hidden")
            hidden.sheet_state = "hidden"
            workbook.save(path)

            visible = PDFService().get_visible_sheets_fast(path)

        self.assertEqual(visible, ["Visible"])

    def test_sheet_name_exact_matching(self):
        """Test sheet matching between new and old files."""
        visible_new = ["Cover", "Part1", "Part2", "OnlyNew"]
        visible_old = ["Cover", "Part1", "Part2", "OnlyOld"]
        
        matching = [s for s in visible_new if s in visible_old]
        missing_in_old = [s for s in visible_new if s not in visible_old]
        missing_in_new = [s for s in visible_old if s not in visible_new]
        
        self.assertEqual(matching, ["Cover", "Part1", "Part2"])
        self.assertEqual(missing_in_old, ["OnlyNew"])
        self.assertEqual(missing_in_new, ["OnlyOld"])


class TestSettingsServiceDocMode(unittest.TestCase):
    """Test SettingsService loads default doc_mode and print_area."""

    def test_default_settings_include_doc_mode(self):
        # Do not let a developer's persisted user_settings.json override defaults.
        missing_settings = os.path.join(os.path.dirname(__file__), "__missing_user_settings__.json")
        with patch.object(SettingsService, "_get_settings_path", return_value=missing_settings):
            settings_service = SettingsService()
            defaults = settings_service.load_settings()

        self.assertIn("doc_mode", defaults)
        self.assertIn("print_area", defaults)
        self.assertEqual(defaults["doc_mode"], config.DOC_MODE_STANDARD_CTTT)
        self.assertEqual(defaults["print_area"], config.PRINT_AREA_STANDARD_CTTT)


class TestDukcOtherFormSheetLogic(unittest.TestCase):
    """Test DOC_MODE_DUKC_OTHER specific logic: only comparing sheet Form (case-insensitive)."""

    def test_form_sheet_matching_case_insensitive(self):
        """Test that sheet 'Form' is found regardless of casing and other sheets are ignored."""
        visible_new = ["Cover", "Form", "Detail", "Note"]
        visible_old = ["Cover", "FORM", "Detail", "Archive"]

        # In DUKC_OTHER mode, find sheet Form case-insensitively
        form_new = next((s for s in visible_new if s.strip().lower() == "form"), None)
        form_old = next((s for s in visible_old if s.strip().lower() == "form"), None)

        self.assertEqual(form_new, "Form")
        self.assertEqual(form_old, "FORM")

        # Ignored sheets
        other_sheets = [s for s in visible_new if s.strip().lower() != "form"]
        self.assertEqual(other_sheets, ["Cover", "Detail", "Note"])

    def test_form_sheet_missing_handling(self):
        """Test handling when Form sheet is missing in new or old file."""
        visible_new_without_form = ["Cover", "Detail"]
        visible_old_with_form = ["Cover", "form"]

        form_new = next((s for s in visible_new_without_form if s.strip().lower() == "form"), None)
        form_old = next((s for s in visible_old_with_form if s.strip().lower() == "form"), None)

        self.assertIsNone(form_new)
        self.assertEqual(form_old, "form")

    def test_wrong_form_mode_returns_clear_error(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            paths = []
            for filename in ("new.xlsx", "old.xlsx"):
                path = os.path.join(temp_dir, filename)
                workbook = Workbook()
                workbook.active.title = "Data"
                workbook.save(path)
                paths.append(path)

            valid, error = ValidationService.validate_document_mode(
                [paths[0]], [paths[1]], config.DOC_MODE_DUKC_OTHER
            )

        self.assertFalse(valid)
        self.assertIn("sheet 'Form'", error)
        self.assertIn("new.xlsx", error)
        self.assertIn("old.xlsx", error)

    def test_form_workbooks_are_rejected_for_dukc_cttt_mode(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            paths = []
            for filename in ("new.xlsx", "old.xlsx"):
                path = os.path.join(temp_dir, filename)
                workbook = Workbook()
                workbook.active.title = "Form"
                workbook.save(path)
                paths.append(path)

            valid, error = ValidationService.validate_document_mode(
                [paths[0]], [paths[1]], config.DOC_MODE_DUKC_CTTT
            )

        self.assertFalse(valid)
        self.assertIn("chọn nhầm loại", error)
        self.assertIn("Tờ Phát Hành DUKC & Khác", error)

    def test_standard_mode_accepts_matching_green_tabs(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            paths = []
            for filename in ("new.xlsx", "old.xlsx"):
                path = os.path.join(temp_dir, filename)
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = "CTTT"
                worksheet.sheet_properties.tabColor = "92D050"
                workbook.save(path)
                paths.append(path)

            valid, error = ValidationService.validate_document_mode(
                [paths[0]], [paths[1]], config.DOC_MODE_STANDARD_CTTT
            )

        self.assertTrue(valid, error)

    def test_standard_mode_accepts_matching_old_sheet_without_green_tab(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            new_path = os.path.join(temp_dir, "new.xlsx")
            old_path = os.path.join(temp_dir, "old.xlsx")

            new_book = Workbook()
            new_sheet = new_book.active
            new_sheet.title = "011 (2)"
            new_sheet.sheet_properties.tabColor = "92D050"
            new_book.save(new_path)

            old_book = Workbook()
            old_sheet = old_book.active
            old_sheet.title = "011 (2)"
            old_sheet.sheet_properties.tabColor = "00CC00"
            old_book.save(old_path)

            valid, error = ValidationService.validate_document_mode(
                [new_path], [old_path], config.DOC_MODE_STANDARD_CTTT
            )

        self.assertTrue(valid, error)

    def test_standard_mode_rejects_new_green_sheet_missing_from_old_file(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            new_path = os.path.join(temp_dir, "new.xlsx")
            old_path = os.path.join(temp_dir, "old.xlsx")

            new_book = Workbook()
            new_sheet = new_book.active
            new_sheet.title = "Selected"
            new_sheet.sheet_properties.tabColor = "92D050"
            new_book.save(new_path)

            old_book = Workbook()
            old_book.active.title = "Different"
            old_book.save(old_path)

            valid, error = ValidationService.validate_document_mode(
                [new_path], [old_path], config.DOC_MODE_STANDARD_CTTT
            )

        self.assertFalse(valid)
        self.assertTrue("trùng tên" in error or "tương ứng" in error)

    def test_form_workbooks_are_rejected_for_standard_mode(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            paths = []
            for filename in ("new.xlsx", "old.xlsx"):
                path = os.path.join(temp_dir, filename)
                workbook = Workbook()
                workbook.active.title = "Form"
                workbook.active.sheet_properties.tabColor = "92D050"
                workbook.save(path)
                paths.append(path)

            valid, error = ValidationService.validate_document_mode(
                [paths[0]], [paths[1]], config.DOC_MODE_STANDARD_CTTT
            )

        self.assertFalse(valid)
        self.assertIn("không phù hợp", error)
        self.assertIn("Tờ Phát Hành DUKC & Khác", error)


class TestPageSetupFitToOnePage(unittest.TestCase):
    """Test that PageSetup sets FitToPagesWide=1 and FitToPagesTall=1 so A1:AT120 does not split."""

    def test_export_sheets_to_pdf_pagesetup(self):
        pdf_service = PDFService()
        pdf_service.excel_app = MagicMock()

        mock_sheet = MagicMock()
        mock_sheet.Name = "Form"
        mock_ps = MagicMock()
        mock_sheet.PageSetup = mock_ps

        mock_wb = MagicMock()
        mock_wb.Sheets = [mock_sheet]
        pdf_service.excel_app.Workbooks.Open.return_value = mock_wb

        with patch("utils.unblock_file"), \
             patch("os.path.exists", return_value=True), \
             patch("shutil.move"), \
             patch("fitz.open") as mock_fitz:
            mock_doc = MagicMock()
            mock_doc.__len__.return_value = 1
            mock_fitz.return_value = mock_doc

            pdf_service.export_sheets_to_pdf(
                "dummy.xlsx", ["Form"], "dummy_out.pdf", print_area="A1:AT120"
            )

            # Verify PageSetup properties
            self.assertEqual(mock_ps.PrintArea, "A1:AT120")
            self.assertEqual(mock_ps.Zoom, False)
            self.assertEqual(mock_ps.FitToPagesWide, 1)
            self.assertEqual(mock_ps.FitToPagesTall, 1)
            self.assertEqual(mock_ps.CenterHorizontally, True)

    def test_sync_sheet_layout_copies_fractional_geometry(self):
        class Dimension:
            def __init__(self, width=None, height=None, hidden=False):
                self.ColumnWidth = width
                self.RowHeight = height
                self.Hidden = hidden
                self.EntireColumn = self
                self.EntireRow = self

        class Dimensions:
            def __init__(self, values):
                self.values = values
                self.Count = len(values)

            def __call__(self, index):
                return self.values[index - 1]

        class Range:
            def __init__(self, columns, rows):
                self.Columns = Dimensions(columns)
                self.Rows = Dimensions(rows)

        class Sheet:
            def __init__(self, area):
                self.area = area

            def Range(self, _address):
                return self.area

        source = Range(
            [Dimension(width=2.453125), Dimension(width=13.7265625, hidden=True)],
            [Dimension(height=14.0), Dimension(height=15.65, hidden=True)],
        )
        target = Range(
            [Dimension(width=2.42578125), Dimension(width=13.7109375)],
            [Dimension(height=15.0), Dimension(height=15.6)],
        )

        PDFService._sync_sheet_layout(Sheet(target), Sheet(source), "A1:B2")

        self.assertEqual(target.Columns(1).ColumnWidth, 2.453125)
        self.assertEqual(target.Columns(2).ColumnWidth, 13.7265625)
        self.assertTrue(target.Columns(2).Hidden)
        self.assertEqual(target.Rows(1).RowHeight, 14.0)
        self.assertEqual(target.Rows(2).RowHeight, 15.65)
        self.assertTrue(target.Rows(2).Hidden)

    def test_layout_dimensions_are_grouped_for_batched_com_writes(self):
        dimensions = [
            {"index": 1, "size": 15.0, "hidden": False},
            {"index": 2, "size": 15.0, "hidden": False},
            {"index": 3, "size": 15.0, "hidden": True},
            {"index": 4, "size": 20.0, "hidden": False},
            {"index": 5, "size": 20.0, "hidden": False},
        ]

        groups = PDFService._group_layout_dimensions(dimensions)

        self.assertEqual(
            [(group["start"], group["end"], group["key"]) for group in groups],
            [
                (1, 2, (15.0, False)),
                (3, 3, (15.0, True)),
                (4, 5, (20.0, False)),
            ],
        )


class TestCustomDocModeLogic(unittest.TestCase):
    """Test validation and custom sheet matching logic for DOC_MODE_CUSTOM."""

    def test_custom_mode_validation_success(self):
        settings = {
            config.KEY_CUSTOM_PRINT_AREA: "A1:Z50",
            config.KEY_CUSTOM_SHEET_MODE: config.CUSTOM_SHEET_MODE_ALL,
        }
        valid, msg = ValidationService.validate_document_mode([], [], config.DOC_MODE_CUSTOM, settings)
        self.assertTrue(valid)
        self.assertIsNone(msg)

    def test_custom_mode_validation_empty_specified_sheets(self):
        settings = {
            config.KEY_CUSTOM_PRINT_AREA: "A1:Z50",
            config.KEY_CUSTOM_SHEET_MODE: config.CUSTOM_SHEET_MODE_SPECIFIED,
            config.KEY_CUSTOM_SPECIFIED_SHEETS: "   ",
        }
        valid, msg = ValidationService.validate_document_mode([], [], config.DOC_MODE_CUSTOM, settings)
        self.assertFalse(valid)
        self.assertIn("vui lòng nhập ít nhất một tên sheet", msg.lower())

    def test_custom_mode_validation_empty_print_area(self):
        settings = {
            config.KEY_CUSTOM_PRINT_AREA: "   ",
            config.KEY_CUSTOM_SHEET_MODE: config.CUSTOM_SHEET_MODE_ALL,
        }
        valid, msg = ValidationService.validate_document_mode([], [], config.DOC_MODE_CUSTOM, settings)
        self.assertFalse(valid)
        self.assertIn("vui lòng nhập phạm vi so sánh", msg.lower())

    def test_settings_service_defaults_for_custom_mode(self):
        service = SettingsService()
        defaults = service.load_settings()
        self.assertEqual(defaults.get(config.KEY_CUSTOM_PRINT_AREA), config.DEFAULT_CUSTOM_PRINT_AREA)
        self.assertEqual(defaults.get(config.KEY_CUSTOM_SHEET_MODE), config.CUSTOM_SHEET_MODE_ALL)
        self.assertEqual(defaults.get(config.KEY_CUSTOM_SPECIFIED_SHEETS), "")
        self.assertEqual(defaults.get(config.KEY_CUSTOM_ONLY_GREEN), False)

    def test_custom_dialog_saves_user_configuration(self):
        dialog = CustomModeConfigDialog.__new__(CustomModeConfigDialog)
        dialog.range_var = MagicMock(get=MagicMock(return_value="b2:p50"))
        dialog.sheet_mode_var = MagicMock(get=MagicMock(return_value=config.CUSTOM_SHEET_MODE_SPECIFIED))
        dialog.specified_sheets_var = MagicMock(get=MagicMock(return_value="Form, Data"))
        dialog.only_green_var = MagicMock(get=MagicMock(return_value=True))
        dialog.settings_service = MagicMock()
        dialog.on_save_callback = MagicMock()
        dialog.destroy = MagicMock()

        dialog._on_save()

        expected = {
            config.KEY_CUSTOM_PRINT_AREA: "B2:P50",
            config.KEY_CUSTOM_SHEET_MODE: config.CUSTOM_SHEET_MODE_SPECIFIED,
            config.KEY_CUSTOM_SPECIFIED_SHEETS: "Form, Data",
            config.KEY_CUSTOM_ONLY_GREEN: True,
        }
        dialog.settings_service.save_settings.assert_called_once_with(expected)
        dialog.on_save_callback.assert_called_once_with(expected)
        dialog.destroy.assert_called_once_with()

    def test_open_custom_dialog_uses_settings_service_and_syncs_saved_area(self):
        app = MainWindow.__new__(MainWindow)
        app.master = MagicMock()
        app.current_lang = "vi"
        app.settings_service = MagicMock()
        app.settings_service.settings = {config.KEY_CUSTOM_PRINT_AREA: "A1:Z100"}
        app.settings = {}
        app.print_area_var = MagicMock()
        app.update_status = MagicMock()

        with patch("ui.main_window_modern.CustomModeConfigDialog") as dialog_class:
            app.open_custom_config_dialog()

        args = dialog_class.call_args.args
        self.assertIs(args[1], app.settings_service)
        self.assertEqual(args[2], "vi")
        callback = args[3]
        callback({config.KEY_CUSTOM_PRINT_AREA: "J2:BD76"})
        app.print_area_var.set.assert_called_once_with("J2:BD76")


if __name__ == "__main__":
    unittest.main()
